#scripts/data/json_to_sqlite.py
#cd D:\shared\TopicClassification\menu-planner
#python scripts/data/json_to_sqlite.py data/mock_menu_dataset.json data/menu.db --export-xlsx data/menu_export.xlsx

#cd /d D:\shared\TopicClassification\menu-planner
#python scripts\json_to_sqlite.py "D:\shared\TopicClassification\GenerativeLanguageModel\czjLLM\LLMSourcePool_Finished\明細" data\menu.db --mode rebuild --auto-stub-missing-ingredients --missing-report data\missing_ingredients.txt --missing-qty-policy skip --bad-links-report data\bad_dish_links.jsonl --missing-price-policy skip --bad-prices-report data\bad_prices.jsonl --missing-inventory-qty-policy zero --bad-inventory-report data\bad_inventory.jsonl --export-xlsx data\menu_export.xlsx
#你要選哪個 policy？
#skip（推薦）：資料可信度較高，但某些菜會缺食材連結
#zero：不缺連結，但 qty=0 可能讓成本/份量計算怪怪的
#one：qty=1 只是「占位」，偏向讓系統能跑、之後再回填

import argparse
import json
import re
import sqlite3
from pathlib import Path
from typing import Dict, Any, List, Iterable, Set, Optional, Union
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

DDL_INIT = """
PRAGMA foreign_keys = ON;

CREATE TABLE IF NOT EXISTS meta (
  key TEXT PRIMARY KEY,
  value TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS ingredients (
  id TEXT PRIMARY KEY,
  name TEXT NOT NULL,
  category TEXT NOT NULL,
  protein_group TEXT,
  default_unit TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS dishes (
  id TEXT PRIMARY KEY,
  name TEXT NOT NULL,
  role TEXT NOT NULL CHECK(role IN ('main','side','veg','soup','fruit')),
  cuisine TEXT,
  meat_type TEXT,
  tags_json TEXT NOT NULL DEFAULT '[]'
);

CREATE TABLE IF NOT EXISTS dish_ingredients (
  dish_id TEXT NOT NULL,
  ingredient_id TEXT NOT NULL,
  qty REAL NOT NULL,
  unit TEXT NOT NULL,
  PRIMARY KEY (dish_id, ingredient_id),
  FOREIGN KEY (dish_id) REFERENCES dishes(id) ON DELETE CASCADE,
  FOREIGN KEY (ingredient_id) REFERENCES ingredients(id) ON DELETE RESTRICT
);

CREATE TABLE IF NOT EXISTS ingredient_prices (
  ingredient_id TEXT NOT NULL,
  price_date TEXT NOT NULL,               -- YYYY-MM-DD
  price_per_unit REAL NOT NULL,
  unit TEXT NOT NULL,
  PRIMARY KEY (ingredient_id, price_date),
  FOREIGN KEY (ingredient_id) REFERENCES ingredients(id) ON DELETE CASCADE
);

CREATE TABLE IF NOT EXISTS inventory (
  ingredient_id TEXT PRIMARY KEY,
  qty_on_hand REAL NOT NULL,
  unit TEXT NOT NULL,
  updated_at TEXT NOT NULL,               -- YYYY-MM-DD
  expiry_date TEXT,                       -- YYYY-MM-DD (nullable)
  FOREIGN KEY (ingredient_id) REFERENCES ingredients(id) ON DELETE CASCADE
);

CREATE TABLE IF NOT EXISTS unit_conversions (
  from_unit TEXT NOT NULL,
  to_unit TEXT NOT NULL,
  factor REAL NOT NULL,
  PRIMARY KEY (from_unit, to_unit)
);

-- Indexes for faster query
CREATE INDEX IF NOT EXISTS idx_dishes_role ON dishes(role);
CREATE INDEX IF NOT EXISTS idx_dishes_meat_type ON dishes(meat_type);
CREATE INDEX IF NOT EXISTS idx_di_ingredient ON dish_ingredients(ingredient_id);
CREATE INDEX IF NOT EXISTS idx_prices_date ON ingredient_prices(price_date);
"""


DDL_REBUILD = """
PRAGMA foreign_keys = ON;

DROP TABLE IF EXISTS unit_conversions;
DROP TABLE IF EXISTS inventory;
DROP TABLE IF EXISTS ingredient_prices;
DROP TABLE IF EXISTS dish_ingredients;
DROP TABLE IF EXISTS dishes;
DROP TABLE IF EXISTS ingredients;
DROP TABLE IF EXISTS meta;
"""


def _as_text(v: Any) -> str:
  if v is None:
    return ""
  if isinstance(v, (dict, list)):
    return json.dumps(v, ensure_ascii=False)
  return str(v)


#def load_json(path: Path) -> Any:
#  with path.open("r", encoding="utf-8") as f:
#    return json.load(f)

def load_json(path: Path) -> Any:
  # utf-8-sig 可吃到 BOM（常見於某些 txt/json 匯出）
  with path.open("r", encoding="utf-8-sig") as f:
    return json.load(f)

def list_input_files(input_path: Path, patterns_csv: str) -> List[Path]:
  """
  input_path 若為資料夾：遞迴搜尋所有符合 patterns_csv（逗號分隔）的檔案。
  預設支援 *.json,*.txt
  """
  pats = [p.strip() for p in (patterns_csv or "").split(",") if p.strip()]
  if not pats:
    pats = ["*.json", "*.txt"]

  files: List[Path] = []
  for pat in pats:
    # ✅ rglob：遞迴搜尋所有子目錄
    files.extend([p for p in input_path.rglob(pat) if p.is_file()])

  # 去重 + 排序（用完整路徑排序，避免不同資料夾同名檔案排序不穩）
  uniq = sorted({p.resolve() for p in files}, key=lambda x: str(x).lower())

  if not uniq:
    raise ValueError(f"目錄下找不到符合 {pats} 的檔案：{input_path}")
  return uniq

def normalize_payloads(raw: Any) -> List[Dict[str, Any]]:
  """
  支援：
  1) 單一 JSON 物件：{...}
  2) 多個 JSON 物件清單：[{...}, {...}]
  """
  if isinstance(raw, dict):
    return [raw]
  if isinstance(raw, list):
    out: List[Dict[str, Any]] = []
    for i, item in enumerate(raw):
      if not isinstance(item, dict):
        raise ValueError(f"JSON 清單第 {i} 筆不是物件（dict）：{type(item)}")
      out.append(item)
    return out
  raise ValueError(f"JSON 根節點必須是 dict 或 list[dict]，但收到：{type(raw)}")


def merge_payloads(payloads: List[Dict[str, Any]]) -> Dict[str, Any]:
  """
  將多個 payload 合併為單一 dataset，避免跨物件引用驗證失敗。
  規則：同一個 key / id 以後面出現的資料覆蓋前面（last wins）。
  """
  merged_meta: Dict[str, Any] = {}

  convs_by_key: Dict[tuple, Dict[str, Any]] = {}
  ingredients_by_id: Dict[str, Dict[str, Any]] = {}
  dishes_by_id: Dict[str, Dict[str, Any]] = {}
  prices_by_key: Dict[tuple, Dict[str, Any]] = {}
  inventory_by_id: Dict[str, Dict[str, Any]] = {}

  for p in payloads:
    # meta：直接合併，後者覆蓋前者
    merged_meta.update(p.get("meta", {}) or {})

    # unit_conversions：以 (from_unit, to_unit) 當 key
    for c in (p.get("unit_conversions", []) or []):
      k = (c.get("from_unit"), c.get("to_unit"))
      convs_by_key[k] = c

    # ingredients：以 id 當 key
    for ing in (p.get("ingredients", []) or []):
      ingredients_by_id[ing["id"]] = ing

    # dishes：以 id 當 key（整筆覆蓋，包含 ingredients 清單）
    for d in (p.get("dishes", []) or []):
      dishes_by_id[d["id"]] = d

    # prices：以 (ingredient_id, price_date) 當 key
    for pr in (p.get("prices", []) or []):
      k = (pr["ingredient_id"], pr["price_date"])
      prices_by_key[k] = pr

    # inventory：以 ingredient_id 當 key
    for inv in (p.get("inventory", []) or []):
      inventory_by_id[inv["ingredient_id"]] = inv

  return {
    "meta": merged_meta,
    "unit_conversions": list(convs_by_key.values()),
    "ingredients": list(ingredients_by_id.values()),
    "dishes": list(dishes_by_id.values()),
    "prices": list(prices_by_key.values()),
    "inventory": list(inventory_by_id.values()),
  }


def as_float(v: Any, default: Optional[float] = None) -> Optional[float]:
  if v is None:
    return default
  if isinstance(v, (int, float)):
    return float(v)
  try:
    s = str(v).strip()
    if not s:
      return default
    s = s.replace(",", "")  # 例如 "1,200"
    return float(s)
  except Exception:
    return default

def clean_id(s: Any) -> str:
  if s is None:
    return ""
  s = str(s)
  # 常見：全形空白、前後空白
  s = s.replace("\u3000", " ").strip()

  # 常見 LLM 手滑：ign_ -> ing_
  if s.startswith("ign_"):
    s = "ing_" + s[4:]

  return s

def write_bad_links_report(path: Path, bad_links: List[Dict[str, Any]]) -> None:
  path.parent.mkdir(parents=True, exist_ok=True)
  # 用 JSONL：一行一筆，最好機器處理
  with path.open("w", encoding="utf-8") as f:
    for r in bad_links:
      f.write(json.dumps(r, ensure_ascii=False) + "\n")
      
def normalize_dataset_ids_inplace(data: Dict[str, Any]) -> None:
  """把 dataset 內所有會當作 key/foreign key 的欄位做 strip/修正"""
  for ing in (data.get("ingredients", []) or []):
    ing["id"] = clean_id(ing.get("id"))

  for d in (data.get("dishes", []) or []):
    d["id"] = clean_id(d.get("id"))
    for di in (d.get("ingredients", []) or []):
      di["ingredient_id"] = clean_id(di.get("ingredient_id"))
      if "unit" in di and di["unit"] is not None:
        di["unit"] = str(di["unit"]).strip()

  for inv in (data.get("inventory", []) or []):
    inv["ingredient_id"] = clean_id(inv.get("ingredient_id"))

  for p in (data.get("prices", []) or []):
    p["ingredient_id"] = clean_id(p.get("ingredient_id"))

  for c in (data.get("unit_conversions", []) or []):
    if "from_unit" in c and c["from_unit"] is not None:
      c["from_unit"] = str(c["from_unit"]).strip()
    if "to_unit" in c and c["to_unit"] is not None:
      c["to_unit"] = str(c["to_unit"]).strip()


def find_missing_ingredient_ids(conn: sqlite3.Connection, data: Dict[str, Any]) -> List[str]:
  ingredients = data.get("ingredients", []) or []
  dishes = data.get("dishes", []) or []
  prices = data.get("prices", []) or []
  inventory = data.get("inventory", []) or []

  ing_ids = {clean_id(x.get("id")) for x in ingredients if x.get("id")}

  needed: Set[str] = set()
  for dish in dishes:
    for di in (dish.get("ingredients", []) or []):
      needed.add(clean_id(di.get("ingredient_id")))

  for inv in inventory:
    needed.add(clean_id(inv.get("ingredient_id")))

  for p in prices:
    needed.add(clean_id(p.get("ingredient_id")))

  needed.discard("")  # 空字串不要算

  missing_from_file = needed - ing_ids
  if not missing_from_file:
    return []

  found_in_db = _existing_ids(conn, "ingredients", "id", missing_from_file)
  missing = sorted(list(missing_from_file - found_in_db))
  return missing


def add_placeholder_ingredients(data: Dict[str, Any], missing_ids: List[str]) -> None:
  if not missing_ids:
    return
  if "ingredients" not in data or data["ingredients"] is None:
    data["ingredients"] = []

  existing = {clean_id(x.get("id")) for x in (data.get("ingredients", []) or []) if x.get("id")}
  for mid in missing_ids:
    mid = clean_id(mid)
    if not mid or mid in existing:
      continue

    # 名稱：盡量取底線後半段（例如 ing_青椒絲 -> 青椒絲、inc_pork_belly -> pork_belly）
    name = mid.split("_", 1)[1] if "_" in mid else mid

    data["ingredients"].append({
      "id": mid,
      "name": name,
      "category": "unknown",
      "protein_group": None,
      "default_unit": "g"
    })
    existing.add(mid)


def write_missing_report(path: Path, missing_ids: List[str]) -> None:
  path.parent.mkdir(parents=True, exist_ok=True)
  path.write_text("\n".join(missing_ids) + "\n", encoding="utf-8")
  

def open_db(db_path: Path) -> sqlite3.Connection:
  conn = sqlite3.connect(str(db_path))
  conn.execute("PRAGMA foreign_keys = ON;")
  return conn


def ensure_schema(conn: sqlite3.Connection) -> None:
  conn.executescript(DDL_INIT)


def rebuild_schema(conn: sqlite3.Connection) -> None:
  conn.executescript(DDL_REBUILD)
  conn.executescript(DDL_INIT)


def _chunked(seq: List[str], size: int = 900) -> Iterable[List[str]]:
  # SQLite 參數數量上限通常 999，留點餘裕
  for i in range(0, len(seq), size):
    yield seq[i:i+size]


def _existing_ids(conn: sqlite3.Connection, table: str, id_col: str, ids: Set[str]) -> Set[str]:
  if not ids:
    return set()
  found: Set[str] = set()
  id_list = list(ids)
  for chunk in _chunked(id_list):
    q = f"SELECT {id_col} FROM {table} WHERE {id_col} IN ({','.join(['?']*len(chunk))})"
    rows = conn.execute(q, chunk).fetchall()
    found.update(r[0] for r in rows)
  return found


def validate_dataset(conn: sqlite3.Connection, data: Dict[str, Any]) -> None:
  # 允許增量 JSON：缺少的區塊視為空
  ingredients = data.get("ingredients", []) or []
  dishes = data.get("dishes", []) or []
  prices = data.get("prices", []) or []
  inventory = data.get("inventory", []) or []
  unit_conversions = data.get("unit_conversions", []) or []

  # 檔案內重複檢查
  ing_ids = set()
  for ing in ingredients:
    if ing["id"] in ing_ids:
      raise ValueError(f"ingredients.id 重複：{ing['id']}")
    ing_ids.add(ing["id"])

  dish_ids = set()
  for dish in dishes:
    if dish["id"] in dish_ids:
      raise ValueError(f"dishes.id 重複：{dish['id']}")
    dish_ids.add(dish["id"])

  # 需要存在的食材 id（可能不在這次 JSON，但可能已在 DB）
  needed_ing_ids: Set[str] = set()

  for dish in dishes:
    for di in dish.get("ingredients", []) or []:
      needed_ing_ids.add(di["ingredient_id"])

  for inv in inventory:
    needed_ing_ids.add(inv["ingredient_id"])

  for p in prices:
    needed_ing_ids.add(p["ingredient_id"])

  # 如果引用的食材不在本檔，就必須已存在 DB
  missing_from_file = needed_ing_ids - ing_ids
  if missing_from_file:
    found_in_db = _existing_ids(conn, "ingredients", "id", missing_from_file)
    missing = missing_from_file - found_in_db
    if missing:
      raise ValueError(f"引用不存在的食材 id（檔案與 DB 都找不到）：{sorted(missing)}")

  # unit_conversions 形式檢查（簡單防呆）
  for c in unit_conversions:
    if c["factor"] is None:
      raise ValueError(f"unit_conversions.factor 不能為空：{c}")


def upsert_meta(conn: sqlite3.Connection, meta: Dict[str, Any]) -> None:
  if not meta:
    return
  rows = [(k, _as_text(v)) for k, v in meta.items()]
  conn.executemany("""
    INSERT INTO meta(key, value) VALUES (?, ?)
    ON CONFLICT(key) DO UPDATE SET value=excluded.value
  """, rows)


def upsert_unit_conversions(conn: sqlite3.Connection, convs: List[Dict[str, Any]]) -> None:
  if not convs:
    return
  rows = [(c["from_unit"], c["to_unit"], float(c["factor"])) for c in convs]
  conn.executemany("""
    INSERT INTO unit_conversions(from_unit, to_unit, factor) VALUES (?, ?, ?)
    ON CONFLICT(from_unit, to_unit) DO UPDATE SET factor=excluded.factor
  """, rows)


def upsert_ingredients(conn: sqlite3.Connection, ingredients: List[Dict[str, Any]]) -> None:
  if not ingredients:
    return

  rows = []
  bad = 0

  for ing in ingredients:
    ing_id = clean_id(ing.get("id"))
    if not ing_id:
      bad += 1
      continue

    name = ing.get("name")
    name = str(name).strip() if name is not None else ""
    if not name:
      # 盡量用 id 後半段當名字
      name = ing_id.split("_", 1)[1] if "_" in ing_id else ing_id

    category = ing.get("category")
    category = str(category).strip() if category is not None else ""
    if not category:
      category = "unknown"

    protein_group = ing.get("protein_group")
    protein_group = str(protein_group).strip() if protein_group is not None else None
    if protein_group == "":
      protein_group = None

    du = ing.get("default_unit")
    du = str(du).strip() if du is not None else ""
    if not du:
      du = "g"  # ✅ 缺漏就補 g，避免 NOT NULL 爆掉

    rows.append((ing_id, name, category, protein_group, du))

  if bad:
    print(f"⚠️ ingredients 有 {bad} 筆缺 id，已略過。")

  conn.executemany("""
    INSERT INTO ingredients(id, name, category, protein_group, default_unit)
    VALUES (?, ?, ?, ?, ?)
    ON CONFLICT(id) DO UPDATE SET
      name=excluded.name,
      category=excluded.category,
      protein_group=excluded.protein_group,
      default_unit=excluded.default_unit
  """, rows)

def upsert_dishes_and_links(
  conn: sqlite3.Connection,
  dishes: List[Dict[str, Any]],
  sync_dish_links: bool = False,
  *,
  missing_qty_policy: str = "skip",
  bad_links: Optional[List[Dict[str, Any]]] = None,
) -> None:
  if not dishes:
    return

  dish_rows = []
  link_rows = []
  dish_ids = []

  allowed_roles = {"main", "side", "veg", "soup", "fruit"}

  for d in dishes:
    dish_id = clean_id(d.get("id"))
    if not dish_id:
      continue
    dish_ids.append(dish_id)

    name = str(d.get("name") or "").strip()
    if not name:
      name = dish_id.split("_", 1)[1] if "_" in dish_id else dish_id

    role = str(d.get("role") or "").strip().lower()
    if role not in allowed_roles:
      # LLM 常吐 snack/drink/other：統一降級到 side
      role = "side"

    cuisine = d.get("cuisine")
    cuisine = str(cuisine).strip() if cuisine is not None else None
    if cuisine == "":
      cuisine = None

    meat_type = d.get("meat_type")
    meat_type = str(meat_type).strip() if meat_type is not None else None
    if meat_type == "":
      meat_type = None

    dish_rows.append((
      dish_id,
      name,
      role,
      cuisine,
      meat_type,
      json.dumps(d.get("tags", []), ensure_ascii=False)
    ))

    for di in (d.get("ingredients", []) or []):
      ing_id = clean_id(di.get("ingredient_id"))
      if not ing_id:
        if bad_links is not None:
          bad_links.append({
            "dish_id": dish_id,
            "ingredient_id": di.get("ingredient_id"),
            "reason": "missing_ingredient_id",
            "raw": di,
          })
        continue

      qty = as_float(di.get("qty"), default=None)
      unit = str(di.get("unit") or "").strip() or "g"

      if qty is None:
        # 記錄問題
        if bad_links is not None:
          bad_links.append({
            "dish_id": dish_id,
            "ingredient_id": ing_id,
            "reason": "missing_or_invalid_qty",
            "raw_qty": di.get("qty"),
            "raw_unit": di.get("unit"),
            "raw": di,
          })

        if missing_qty_policy == "skip":
          continue
        elif missing_qty_policy == "one":
          qty = 1.0
        else:
          qty = 0.0  # zero

      link_rows.append((dish_id, ing_id, float(qty), unit))

  conn.executemany("""
    INSERT INTO dishes(id, name, role, cuisine, meat_type, tags_json)
    VALUES (?, ?, ?, ?, ?, ?)
    ON CONFLICT(id) DO UPDATE SET
      name=excluded.name,
      role=excluded.role,
      cuisine=excluded.cuisine,
      meat_type=excluded.meat_type,
      tags_json=excluded.tags_json
  """, dish_rows)

  if sync_dish_links and dish_ids:
    conn.executemany("DELETE FROM dish_ingredients WHERE dish_id = ?", [(i,) for i in dish_ids])

  if link_rows:
    conn.executemany("""
      INSERT INTO dish_ingredients(dish_id, ingredient_id, qty, unit)
      VALUES (?, ?, ?, ?)
      ON CONFLICT(dish_id, ingredient_id) DO UPDATE SET
        qty=excluded.qty,
        unit=excluded.unit
    """, link_rows)


def upsert_prices(
  conn: sqlite3.Connection,
  prices: List[Dict[str, Any]],
  *,
  missing_price_policy: str = "skip",
  bad_prices: Optional[List[Dict[str, Any]]] = None,
) -> None:
  if not prices:
    return

  rows = []
  for p in prices:
    ing_id = clean_id(p.get("ingredient_id"))
    price_date = str(p.get("price_date") or "").strip()
    unit = str(p.get("unit") or "").strip() or "g"

    v = as_float(p.get("price_per_unit"), default=None)

    # 基本必填欄位檢查（price_date / ingredient_id）
    if not ing_id or not price_date:
      if bad_prices is not None:
        bad_prices.append({
          "ingredient_id": p.get("ingredient_id"),
          "price_date": p.get("price_date"),
          "reason": "missing_ingredient_id_or_price_date",
          "raw": p,
        })
      continue

    if v is None:
      if bad_prices is not None:
        bad_prices.append({
          "ingredient_id": ing_id,
          "price_date": price_date,
          "reason": "missing_or_invalid_price_per_unit",
          "raw_price_per_unit": p.get("price_per_unit"),
          "raw_unit": p.get("unit"),
          "raw": p,
        })

      if missing_price_policy == "skip":
        continue
      v = 0.0

    rows.append((ing_id, price_date, float(v), unit))

  if not rows:
    return

  conn.executemany("""
    INSERT INTO ingredient_prices(ingredient_id, price_date, price_per_unit, unit)
    VALUES (?, ?, ?, ?)
    ON CONFLICT(ingredient_id, price_date) DO UPDATE SET
      price_per_unit=excluded.price_per_unit,
      unit=excluded.unit
  """, rows)


def upsert_inventory(
  conn: sqlite3.Connection,
  inv_list: List[Dict[str, Any]],
  *,
  missing_qty_policy: str = "zero",
  bad_inventory: Optional[List[Dict[str, Any]]] = None,
) -> None:
  if not inv_list:
    return

  rows = []
  today = datetime.now().strftime("%Y-%m-%d")

  for inv in inv_list:
    ing_id = clean_id(inv.get("ingredient_id"))
    unit = str(inv.get("unit") or "").strip() or "g"
    updated_at = str(inv.get("updated_at") or "").strip() or today
    expiry_date = inv.get("expiry_date")

    qty = as_float(inv.get("qty_on_hand"), default=None)

    if not ing_id:
      if bad_inventory is not None:
        bad_inventory.append({
          "ingredient_id": inv.get("ingredient_id"),
          "reason": "missing_ingredient_id",
          "raw": inv,
        })
      continue

    if qty is None:
      if bad_inventory is not None:
        bad_inventory.append({
          "ingredient_id": ing_id,
          "reason": "missing_or_invalid_qty_on_hand",
          "raw_qty_on_hand": inv.get("qty_on_hand"),
          "raw_unit": inv.get("unit"),
          "raw": inv,
        })
      if missing_qty_policy == "skip":
        continue
      qty = 0.0

    rows.append((ing_id, float(qty), unit, updated_at, expiry_date))

  if not rows:
    return

  conn.executemany("""
    INSERT INTO inventory(ingredient_id, qty_on_hand, unit, updated_at, expiry_date)
    VALUES (?, ?, ?, ?, ?)
    ON CONFLICT(ingredient_id) DO UPDATE SET
      qty_on_hand=excluded.qty_on_hand,
      unit=excluded.unit,
      updated_at=excluded.updated_at,
      expiry_date=excluded.expiry_date
  """, rows)

def _safe_sheet_title(title: str) -> str:
  # Excel 工作表名稱限制：不可含 \ / ? * [ ] : ，且最長 31
  bad = ['\\', '/', '?', '*', '[', ']', ':']
  for ch in bad:
    title = title.replace(ch, '_')
  title = title.strip() or "sheet"
  return title[:31]


def _fetch_all(conn: sqlite3.Connection, sql: str, params: tuple = ()) -> tuple[list[str], list[tuple]]:
  cur = conn.execute(sql, params)
  cols = [d[0] for d in cur.description]
  rows = cur.fetchall()
  return cols, rows


def _write_sheet(wb: Workbook, sheet_name: str, cols: list[str], rows: list[tuple]) -> None:
  ws = wb.create_sheet(title=_safe_sheet_title(sheet_name))
  ws.append(cols)
  for r in rows:
    ws.append(list(r))

  # 凍結首列、加上篩選
  ws.freeze_panes = "A2"
  if cols:
    last_col = get_column_letter(len(cols))
    last_row = max(1, len(rows) + 1)
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"

  # 簡單自動欄寬（避免太慢：只掃前 200 筆）
  sample_n = min(200, len(rows))
  for j, col in enumerate(cols, start=1):
    max_len = len(str(col))
    for i in range(sample_n):
      v = rows[i][j - 1]
      if v is None:
        continue
      max_len = max(max_len, len(str(v)))
    ws.column_dimensions[get_column_letter(j)].width = min(60, max(10, max_len + 2))


def export_db_to_excel(conn: sqlite3.Connection, xlsx_path: Path) -> None:
  wb = Workbook()
  # 預設會有一張空白工作表，先移除
  default_ws = wb.active
  wb.remove(default_ws)

  table_queries = [
    ("meta", "SELECT key, value FROM meta ORDER BY key"),
    ("ingredients", "SELECT id, name, category, protein_group, default_unit FROM ingredients ORDER BY id"),
    ("dishes", "SELECT id, name, role, cuisine, meat_type, tags_json FROM dishes ORDER BY role, id"),
    ("dish_ingredients", "SELECT dish_id, ingredient_id, qty, unit FROM dish_ingredients ORDER BY dish_id, ingredient_id"),
    ("ingredient_prices", "SELECT ingredient_id, price_date, price_per_unit, unit FROM ingredient_prices ORDER BY price_date, ingredient_id"),
    ("inventory", "SELECT ingredient_id, qty_on_hand, unit, updated_at, expiry_date FROM inventory ORDER BY ingredient_id"),
    ("unit_conversions", "SELECT from_unit, to_unit, factor FROM unit_conversions ORDER BY from_unit, to_unit"),
  ]

  for sheet_name, sql in table_queries:
    cols, rows = _fetch_all(conn, sql)
    _write_sheet(wb, sheet_name, cols, rows)

  xlsx_path.parent.mkdir(parents=True, exist_ok=True)
  wb.save(str(xlsx_path))
  
def main():
  parser = argparse.ArgumentParser(description="JSON → SQLite（支援增量匯入）")
  parser.add_argument("--input-glob", type=str, default="*.json,*.txt",
                      help="當 input_json 是資料夾時，用此 glob 選檔（逗號分隔），例如：\"*.txt\" 或 \"*.json,*.txt\"")
  parser.add_argument("input_json", type=str, help="輸入 JSON 檔")
  parser.add_argument("output_db", type=str, help="輸出 / 目標 SQLite DB 檔")
  parser.add_argument("--mode", choices=["upsert", "rebuild"], default="upsert",
                      help="upsert：保留舊資料、增量新增/更新；rebuild：重建並清空資料")
  parser.add_argument("--sync-dish-links", action="store_true",
                      help="以本次 JSON 的 dish.ingredients 覆蓋該 dish 的食材連結（會先刪再寫）")
  parser.add_argument("--export-xlsx", type=str, default="",
                      help="匯入完成後輸出 Excel（.xlsx），每張表一個工作表")
  parser.add_argument("--auto-stub-missing-ingredients", action="store_true",
                      help="遇到引用不存在的 ingredient_id 時，自動補 placeholder 食材（category=unknown, default_unit=g）")
  parser.add_argument("--missing-report", type=str, default="",
                      help="把缺漏食材 id 輸出成文字檔（一行一個 id）")
  parser.add_argument("--missing-qty-policy", choices=["skip", "zero", "one"], default="skip",
                    help="dish_ingredients.qty 缺漏時的處理：skip=略過該連結；zero=補 0；one=補 1")
  parser.add_argument("--bad-links-report", type=str, default="",
                    help="輸出 dish_ingredients 異常連結清單（JSONL）")
  parser.add_argument("--missing-price-policy", choices=["skip", "zero"], default="skip",
                      help="ingredient_prices.price_per_unit 缺漏時：skip=略過該價格；zero=補 0")
  parser.add_argument("--bad-prices-report", type=str, default="",
                      help="輸出 ingredient_prices 異常清單（JSONL）")

  parser.add_argument("--missing-inventory-qty-policy", choices=["skip", "zero"], default="zero",
                      help="inventory.qty_on_hand 缺漏時：skip=略過該庫存；zero=補 0")
  parser.add_argument("--bad-inventory-report", type=str, default="",
                      help="輸出 inventory 異常清單（JSONL）")

  args = parser.parse_args()

  json_path = Path(args.input_json).resolve()
  db_path = Path(args.output_db).resolve()

  input_path = Path(args.input_json).resolve()

  # input_json 可為「檔案」或「資料夾」
  if input_path.is_dir():
    input_files = list_input_files(input_path, args.input_glob)
  else:
    input_files = [input_path]

  payloads: List[Dict[str, Any]] = []
  for fp in input_files:
    raw = load_json(fp)
    payloads.extend(normalize_payloads(raw))

  data = merge_payloads(payloads)

  # ✅ 先清理 id（修掉前後空白、ign_ 等）
  normalize_dataset_ids_inplace(data)

  conn = open_db(db_path)
  try:
    with conn:
      if args.mode == "rebuild":
        rebuild_schema(conn)
      else:
        ensure_schema(conn)

      # ✅ 找缺漏食材
      missing_ids = find_missing_ingredient_ids(conn, data)
      if missing_ids:
        if args.auto_stub_missing_ingredients:
          add_placeholder_ingredients(data, missing_ids)
          print(f"⚠️ 缺漏食材 id：{len(missing_ids)} 筆，已自動補 placeholder。")
          if args.missing_report:
            rp = Path(args.missing_report).resolve()
            write_missing_report(rp, missing_ids)
            print(f"🧾 缺漏清單已輸出：{rp}")
        else:
          # 保持原本嚴格模式：直接讓 validate_dataset 擋掉
          pass

      # 合併後再驗證（若開了 auto-stub，這裡通常會通過）
      validate_dataset(conn, data)

      # 寫入
      upsert_meta(conn, data.get("meta", {}) or {})
      upsert_unit_conversions(conn, data.get("unit_conversions", []) or [])
      upsert_ingredients(conn, data.get("ingredients", []) or [])
      bad_links: List[Dict[str, Any]] = []
    
      upsert_dishes_and_links(
        conn,
        data.get("dishes", []) or [],
        sync_dish_links=args.sync_dish_links,
        missing_qty_policy=args.missing_qty_policy,
        bad_links=bad_links,
      )
    
      if args.bad_links_report and bad_links:
        rp = Path(args.bad_links_report).resolve()
        write_bad_links_report(rp, bad_links)
        print(f"🧾 異常連結清單已輸出：{rp}（{len(bad_links)} 筆）")
      bad_prices: List[Dict[str, Any]] = []
      upsert_prices(
        conn,
        data.get("prices", []) or [],
        missing_price_policy=args.missing_price_policy,
        bad_prices=bad_prices,
      )
      if args.bad_prices_report and bad_prices:
        rp = Path(args.bad_prices_report).resolve()
        write_bad_links_report(rp, bad_prices)  # 你這個函式其實就是通用 JSONL writer
        print(f"🧾 價格異常清單已輸出：{rp}（{len(bad_prices)} 筆）")

      bad_inventory: List[Dict[str, Any]] = []
      upsert_inventory(
        conn,
        data.get("inventory", []) or [],
        missing_qty_policy=args.missing_inventory_qty_policy,
        bad_inventory=bad_inventory,
      )
      if args.bad_inventory_report and bad_inventory:
        rp = Path(args.bad_inventory_report).resolve()
        write_bad_links_report(rp, bad_inventory)
        print(f"🧾 庫存異常清單已輸出：{rp}（{len(bad_inventory)} 筆）")

      upsert_meta(conn, {"last_import_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")})

    print(f"✅ 匯入完成：{db_path}")

    if args.export_xlsx:
      export_db_to_excel(conn, Path(args.export_xlsx).resolve())
      print(f"📄 已輸出 Excel：{Path(args.export_xlsx).resolve()}")
      
    cur = conn.cursor()
    cur.execute("SELECT role, COUNT(*) FROM dishes GROUP BY role ORDER BY role")
    print("— dishes by role —")
    for r, c in cur.fetchall():
      print(f"{r}: {c}")

    cur.execute("SELECT COUNT(*) FROM dish_ingredients")
    print(f"— dish_ingredients rows — {cur.fetchone()[0]}")

  finally:
    conn.close()


if __name__ == "__main__":
  main()
