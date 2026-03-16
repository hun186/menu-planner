#scripts/data/json_to_sqlite.py
#cd D:\shared\TopicClassification\menu-planner
#python scripts/data/json_to_sqlite.py data/mock_menu_dataset.json data/menu.db --export-xlsx data/menu_export.xlsx
import argparse
import json
import sqlite3
from pathlib import Path
from typing import Dict, Any, List, Iterable, Set, Optional, Union
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

DDL_INIT = """
PRAGMA foreign_keys = ON;

CREATE TABLE IF NOT EXISTS meta (
  key TEXT PRIMARY KEY,
  value TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS ingredients (
  id TEXT PRIMARY KEY,
  name TEXT NOT NULL,
  category TEXT NOT NULL,
  protein_group TEXT,
  default_unit TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS dishes (
  id TEXT PRIMARY KEY,
  name TEXT NOT NULL,
  role TEXT NOT NULL CHECK(role IN ('main','side','soup','fruit')),
  cuisine TEXT,
  meat_type TEXT,
  tags_json TEXT NOT NULL DEFAULT '[]'
);

CREATE TABLE IF NOT EXISTS dish_ingredients (
  dish_id TEXT NOT NULL,
  ingredient_id TEXT NOT NULL,
  qty REAL NOT NULL,
  unit TEXT NOT NULL,
  PRIMARY KEY (dish_id, ingredient_id),
  FOREIGN KEY (dish_id) REFERENCES dishes(id) ON DELETE CASCADE,
  FOREIGN KEY (ingredient_id) REFERENCES ingredients(id) ON DELETE RESTRICT
);

CREATE TABLE IF NOT EXISTS ingredient_prices (
  ingredient_id TEXT NOT NULL,
  price_date TEXT NOT NULL,               -- YYYY-MM-DD
  price_per_unit REAL NOT NULL,
  unit TEXT NOT NULL,
  PRIMARY KEY (ingredient_id, price_date),
  FOREIGN KEY (ingredient_id) REFERENCES ingredients(id) ON DELETE CASCADE
);

CREATE TABLE IF NOT EXISTS inventory (
  ingredient_id TEXT PRIMARY KEY,
  qty_on_hand REAL NOT NULL,
  unit TEXT NOT NULL,
  updated_at TEXT NOT NULL,               -- YYYY-MM-DD
  expiry_date TEXT,                       -- YYYY-MM-DD (nullable)
  FOREIGN KEY (ingredient_id) REFERENCES ingredients(id) ON DELETE CASCADE
);

CREATE TABLE IF NOT EXISTS unit_conversions (
  from_unit TEXT NOT NULL,
  to_unit TEXT NOT NULL,
  factor REAL NOT NULL,
  PRIMARY KEY (from_unit, to_unit)
);

-- Indexes for faster query
CREATE INDEX IF NOT EXISTS idx_dishes_role ON dishes(role);
CREATE INDEX IF NOT EXISTS idx_dishes_meat_type ON dishes(meat_type);
CREATE INDEX IF NOT EXISTS idx_di_ingredient ON dish_ingredients(ingredient_id);
CREATE INDEX IF NOT EXISTS idx_prices_date ON ingredient_prices(price_date);
"""


DDL_REBUILD = """
PRAGMA foreign_keys = ON;

DROP TABLE IF EXISTS unit_conversions;
DROP TABLE IF EXISTS inventory;
DROP TABLE IF EXISTS ingredient_prices;
DROP TABLE IF EXISTS dish_ingredients;
DROP TABLE IF EXISTS dishes;
DROP TABLE IF EXISTS ingredients;
DROP TABLE IF EXISTS meta;
"""


def _as_text(v: Any) -> str:
  if v is None:
    return ""
  if isinstance(v, (dict, list)):
    return json.dumps(v, ensure_ascii=False)
  return str(v)


#def load_json(path: Path) -> Any:
#  with path.open("r", encoding="utf-8") as f:
#    return json.load(f)

def load_json(path: Path) -> Any:
  # utf-8-sig 可吃到 BOM（常見於某些 txt/json 匯出）
  with path.open("r", encoding="utf-8-sig") as f:
    return json.load(f)

def list_input_files(input_path: Path, patterns_csv: str) -> List[Path]:
  """
  input_path 若為資料夾，依 patterns_csv（逗號分隔）列出所有檔案。
  預設支援 *.json,*.txt
  """
  pats = [p.strip() for p in (patterns_csv or "").split(",") if p.strip()]
  if not pats:
    pats = ["*.json", "*.txt"]

  files: List[Path] = []
  for pat in pats:
    files.extend([p for p in input_path.glob(pat) if p.is_file()])

  # 去重 + 排序（用檔名排序，穩定）
  uniq = sorted({p.resolve() for p in files}, key=lambda x: x.name.lower())

  if not uniq:
    raise ValueError(f"資料夾內找不到符合 {pats} 的檔案：{input_path}")
  return uniq

def normalize_payloads(raw: Any) -> List[Dict[str, Any]]:
  """
  支援：
  1) 單一 JSON 物件：{...}
  2) 多個 JSON 物件清單：[{...}, {...}]
  """
  if isinstance(raw, dict):
    return [raw]
  if isinstance(raw, list):
    out: List[Dict[str, Any]] = []
    for i, item in enumerate(raw):
      if not isinstance(item, dict):
        raise ValueError(f"JSON 清單第 {i} 筆不是物件（dict）：{type(item)}")
      out.append(item)
    return out
  raise ValueError(f"JSON 根節點必須是 dict 或 list[dict]，但收到：{type(raw)}")


def merge_payloads(payloads: List[Dict[str, Any]]) -> Dict[str, Any]:
  """
  將多個 payload 合併為單一 dataset，避免跨物件引用驗證失敗。
  規則：同一個 key / id 以後面出現的資料覆蓋前面（last wins）。
  """
  merged_meta: Dict[str, Any] = {}

  convs_by_key: Dict[tuple, Dict[str, Any]] = {}
  ingredients_by_id: Dict[str, Dict[str, Any]] = {}
  dishes_by_id: Dict[str, Dict[str, Any]] = {}
  prices_by_key: Dict[tuple, Dict[str, Any]] = {}
  inventory_by_id: Dict[str, Dict[str, Any]] = {}

  for p in payloads:
    # meta：直接合併，後者覆蓋前者
    merged_meta.update(p.get("meta", {}) or {})

    # unit_conversions：以 (from_unit, to_unit) 當 key
    for c in (p.get("unit_conversions", []) or []):
      k = (c.get("from_unit"), c.get("to_unit"))
      convs_by_key[k] = c

    # ingredients：以 id 當 key
    for ing in (p.get("ingredients", []) or []):
      ingredients_by_id[ing["id"]] = ing

    # dishes：以 id 當 key（整筆覆蓋，包含 ingredients 清單）
    for d in (p.get("dishes", []) or []):
      dishes_by_id[d["id"]] = d

    # prices：以 (ingredient_id, price_date) 當 key
    for pr in (p.get("prices", []) or []):
      k = (pr["ingredient_id"], pr["price_date"])
      prices_by_key[k] = pr

    # inventory：以 ingredient_id 當 key
    for inv in (p.get("inventory", []) or []):
      inventory_by_id[inv["ingredient_id"]] = inv

  return {
    "meta": merged_meta,
    "unit_conversions": list(convs_by_key.values()),
    "ingredients": list(ingredients_by_id.values()),
    "dishes": list(dishes_by_id.values()),
    "prices": list(prices_by_key.values()),
    "inventory": list(inventory_by_id.values()),
  }


def open_db(db_path: Path) -> sqlite3.Connection:
  conn = sqlite3.connect(str(db_path))
  conn.execute("PRAGMA foreign_keys = ON;")
  return conn


def ensure_schema(conn: sqlite3.Connection) -> None:
  conn.executescript(DDL_INIT)


def rebuild_schema(conn: sqlite3.Connection) -> None:
  conn.executescript(DDL_REBUILD)
  conn.executescript(DDL_INIT)


def _chunked(seq: List[str], size: int = 900) -> Iterable[List[str]]:
  # SQLite 參數數量上限通常 999，留點餘裕
  for i in range(0, len(seq), size):
    yield seq[i:i+size]


def _existing_ids(conn: sqlite3.Connection, table: str, id_col: str, ids: Set[str]) -> Set[str]:
  if not ids:
    return set()
  found: Set[str] = set()
  id_list = list(ids)
  for chunk in _chunked(id_list):
    q = f"SELECT {id_col} FROM {table} WHERE {id_col} IN ({','.join(['?']*len(chunk))})"
    rows = conn.execute(q, chunk).fetchall()
    found.update(r[0] for r in rows)
  return found


def validate_dataset(conn: sqlite3.Connection, data: Dict[str, Any]) -> None:
  # 允許增量 JSON：缺少的區塊視為空
  ingredients = data.get("ingredients", []) or []
  dishes = data.get("dishes", []) or []
  prices = data.get("prices", []) or []
  inventory = data.get("inventory", []) or []
  unit_conversions = data.get("unit_conversions", []) or []

  # 檔案內重複檢查
  ing_ids = set()
  for ing in ingredients:
    if ing["id"] in ing_ids:
      raise ValueError(f"ingredients.id 重複：{ing['id']}")
    ing_ids.add(ing["id"])

  dish_ids = set()
  for dish in dishes:
    if dish["id"] in dish_ids:
      raise ValueError(f"dishes.id 重複：{dish['id']}")
    dish_ids.add(dish["id"])

  # 需要存在的食材 id（可能不在這次 JSON，但可能已在 DB）
  needed_ing_ids: Set[str] = set()

  for dish in dishes:
    for di in dish.get("ingredients", []) or []:
      needed_ing_ids.add(di["ingredient_id"])

  for inv in inventory:
    needed_ing_ids.add(inv["ingredient_id"])

  for p in prices:
    needed_ing_ids.add(p["ingredient_id"])

  # 如果引用的食材不在本檔，就必須已存在 DB
  missing_from_file = needed_ing_ids - ing_ids
  if missing_from_file:
    found_in_db = _existing_ids(conn, "ingredients", "id", missing_from_file)
    missing = missing_from_file - found_in_db
    if missing:
      raise ValueError(f"引用不存在的食材 id（檔案與 DB 都找不到）：{sorted(missing)}")

  # unit_conversions 形式檢查（簡單防呆）
  for c in unit_conversions:
    if c["factor"] is None:
      raise ValueError(f"unit_conversions.factor 不能為空：{c}")


def upsert_meta(conn: sqlite3.Connection, meta: Dict[str, Any]) -> None:
  if not meta:
    return
  rows = [(k, _as_text(v)) for k, v in meta.items()]
  conn.executemany("""
    INSERT INTO meta(key, value) VALUES (?, ?)
    ON CONFLICT(key) DO UPDATE SET value=excluded.value
  """, rows)


def upsert_unit_conversions(conn: sqlite3.Connection, convs: List[Dict[str, Any]]) -> None:
  if not convs:
    return
  rows = [(c["from_unit"], c["to_unit"], float(c["factor"])) for c in convs]
  conn.executemany("""
    INSERT INTO unit_conversions(from_unit, to_unit, factor) VALUES (?, ?, ?)
    ON CONFLICT(from_unit, to_unit) DO UPDATE SET factor=excluded.factor
  """, rows)


def upsert_ingredients(conn: sqlite3.Connection, ingredients: List[Dict[str, Any]]) -> None:
  if not ingredients:
    return
  rows = []
  for ing in ingredients:
    rows.append((
      ing["id"],
      ing["name"],
      ing["category"],
      ing.get("protein_group"),
      ing["default_unit"]
    ))
  conn.executemany("""
    INSERT INTO ingredients(id, name, category, protein_group, default_unit)
    VALUES (?, ?, ?, ?, ?)
    ON CONFLICT(id) DO UPDATE SET
      name=excluded.name,
      category=excluded.category,
      protein_group=excluded.protein_group,
      default_unit=excluded.default_unit
  """, rows)


def upsert_dishes_and_links(
  conn: sqlite3.Connection,
  dishes: List[Dict[str, Any]],
  sync_dish_links: bool = False
) -> None:
  if not dishes:
    return

  dish_rows = []
  link_rows = []
  dish_ids = []

  for d in dishes:
    dish_ids.append(d["id"])
    dish_rows.append((
      d["id"],
      d["name"],
      d["role"],
      d.get("cuisine"),
      d.get("meat_type"),
      json.dumps(d.get("tags", []), ensure_ascii=False)
    ))
    for di in d.get("ingredients", []) or []:
      link_rows.append((
        d["id"],
        di["ingredient_id"],
        float(di["qty"]),
        di["unit"]
      ))

  conn.executemany("""
    INSERT INTO dishes(id, name, role, cuisine, meat_type, tags_json)
    VALUES (?, ?, ?, ?, ?, ?)
    ON CONFLICT(id) DO UPDATE SET
      name=excluded.name,
      role=excluded.role,
      cuisine=excluded.cuisine,
      meat_type=excluded.meat_type,
      tags_json=excluded.tags_json
  """, dish_rows)

  if sync_dish_links and dish_ids:
    conn.executemany("DELETE FROM dish_ingredients WHERE dish_id = ?", [(i,) for i in dish_ids])

  if link_rows:
    conn.executemany("""
      INSERT INTO dish_ingredients(dish_id, ingredient_id, qty, unit)
      VALUES (?, ?, ?, ?)
      ON CONFLICT(dish_id, ingredient_id) DO UPDATE SET
        qty=excluded.qty,
        unit=excluded.unit
    """, link_rows)


def upsert_prices(conn: sqlite3.Connection, prices: List[Dict[str, Any]]) -> None:
  if not prices:
    return
  rows = []
  for p in prices:
    rows.append((
      p["ingredient_id"],
      p["price_date"],
      float(p["price_per_unit"]),
      p["unit"]
    ))
  conn.executemany("""
    INSERT INTO ingredient_prices(ingredient_id, price_date, price_per_unit, unit)
    VALUES (?, ?, ?, ?)
    ON CONFLICT(ingredient_id, price_date) DO UPDATE SET
      price_per_unit=excluded.price_per_unit,
      unit=excluded.unit
  """, rows)


def upsert_inventory(conn: sqlite3.Connection, inv_list: List[Dict[str, Any]]) -> None:
  if not inv_list:
    return
  rows = []
  for inv in inv_list:
    rows.append((
      inv["ingredient_id"],
      float(inv["qty_on_hand"]),
      inv["unit"],
      inv["updated_at"],
      inv.get("expiry_date")
    ))
  conn.executemany("""
    INSERT INTO inventory(ingredient_id, qty_on_hand, unit, updated_at, expiry_date)
    VALUES (?, ?, ?, ?, ?)
    ON CONFLICT(ingredient_id) DO UPDATE SET
      qty_on_hand=excluded.qty_on_hand,
      unit=excluded.unit,
      updated_at=excluded.updated_at,
      expiry_date=excluded.expiry_date
  """, rows)

def _safe_sheet_title(title: str) -> str:
  # Excel 工作表名稱限制：不可含 \ / ? * [ ] : ，且最長 31
  bad = ['\\', '/', '?', '*', '[', ']', ':']
  for ch in bad:
    title = title.replace(ch, '_')
  title = title.strip() or "sheet"
  return title[:31]


def _fetch_all(conn: sqlite3.Connection, sql: str, params: tuple = ()) -> tuple[list[str], list[tuple]]:
  cur = conn.execute(sql, params)
  cols = [d[0] for d in cur.description]
  rows = cur.fetchall()
  return cols, rows


def _write_sheet(wb: Workbook, sheet_name: str, cols: list[str], rows: list[tuple]) -> None:
  ws = wb.create_sheet(title=_safe_sheet_title(sheet_name))
  ws.append(cols)
  for r in rows:
    ws.append(list(r))

  # 凍結首列、加上篩選
  ws.freeze_panes = "A2"
  if cols:
    last_col = get_column_letter(len(cols))
    last_row = max(1, len(rows) + 1)
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"

  # 簡單自動欄寬（避免太慢：只掃前 200 筆）
  sample_n = min(200, len(rows))
  for j, col in enumerate(cols, start=1):
    max_len = len(str(col))
    for i in range(sample_n):
      v = rows[i][j - 1]
      if v is None:
        continue
      max_len = max(max_len, len(str(v)))
    ws.column_dimensions[get_column_letter(j)].width = min(60, max(10, max_len + 2))


def export_db_to_excel(conn: sqlite3.Connection, xlsx_path: Path) -> None:
  wb = Workbook()
  # 預設會有一張空白工作表，先移除
  default_ws = wb.active
  wb.remove(default_ws)

  table_queries = [
    ("meta", "SELECT key, value FROM meta ORDER BY key"),
    ("ingredients", "SELECT id, name, category, protein_group, default_unit FROM ingredients ORDER BY id"),
    ("dishes", "SELECT id, name, role, cuisine, meat_type, tags_json FROM dishes ORDER BY role, id"),
    ("dish_ingredients", "SELECT dish_id, ingredient_id, qty, unit FROM dish_ingredients ORDER BY dish_id, ingredient_id"),
    ("ingredient_prices", "SELECT ingredient_id, price_date, price_per_unit, unit FROM ingredient_prices ORDER BY price_date, ingredient_id"),
    ("inventory", "SELECT ingredient_id, qty_on_hand, unit, updated_at, expiry_date FROM inventory ORDER BY ingredient_id"),
    ("unit_conversions", "SELECT from_unit, to_unit, factor FROM unit_conversions ORDER BY from_unit, to_unit"),
  ]

  for sheet_name, sql in table_queries:
    cols, rows = _fetch_all(conn, sql)
    _write_sheet(wb, sheet_name, cols, rows)

  xlsx_path.parent.mkdir(parents=True, exist_ok=True)
  wb.save(str(xlsx_path))
  
def main():
  parser = argparse.ArgumentParser(description="JSON → SQLite（支援增量匯入）")
  parser.add_argument("--input-glob", type=str, default="*.json,*.txt",
                      help="當 input_json 是資料夾時，用此 glob 選檔（逗號分隔），例如：\"*.txt\" 或 \"*.json,*.txt\"")
  parser.add_argument("input_json", type=str, help="輸入 JSON 檔")
  parser.add_argument("output_db", type=str, help="輸出 / 目標 SQLite DB 檔")
  parser.add_argument("--mode", choices=["upsert", "rebuild"], default="upsert",
                      help="upsert：保留舊資料、增量新增/更新；rebuild：重建並清空資料")
  parser.add_argument("--sync-dish-links", action="store_true",
                      help="以本次 JSON 的 dish.ingredients 覆蓋該 dish 的食材連結（會先刪再寫）")
  parser.add_argument("--export-xlsx", type=str, default="",
                      help="匯入完成後輸出 Excel（.xlsx），每張表一個工作表")
  args = parser.parse_args()

  json_path = Path(args.input_json).resolve()
  db_path = Path(args.output_db).resolve()

  input_path = Path(args.input_json).resolve()

  # input_json 可為「檔案」或「資料夾」
  if input_path.is_dir():
    input_files = list_input_files(input_path, args.input_glob)
  else:
    input_files = [input_path]

  payloads: List[Dict[str, Any]] = []
  for fp in input_files:
    raw = load_json(fp)
    payloads.extend(normalize_payloads(raw))

  data = merge_payloads(payloads)

  conn = open_db(db_path)
  try:
    with conn:
      if args.mode == "rebuild":
        rebuild_schema(conn)
      else:
        ensure_schema(conn)

      # 合併後再驗證：支援「分散在多個物件」的引用關係
      validate_dataset(conn, data)

      # 寫入
      upsert_meta(conn, data.get("meta", {}) or {})
      upsert_unit_conversions(conn, data.get("unit_conversions", []) or [])
      upsert_ingredients(conn, data.get("ingredients", []) or [])
      upsert_dishes_and_links(conn, data.get("dishes", []) or [], sync_dish_links=args.sync_dish_links)
      upsert_prices(conn, data.get("prices", []) or [])
      upsert_inventory(conn, data.get("inventory", []) or [])

      upsert_meta(conn, {"last_import_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")})

    print(f"✅ 匯入完成：{db_path}")

    if args.export_xlsx:
      export_db_to_excel(conn, Path(args.export_xlsx).resolve())
      print(f"📄 已輸出 Excel：{Path(args.export_xlsx).resolve()}")
      
    cur = conn.cursor()
    cur.execute("SELECT role, COUNT(*) FROM dishes GROUP BY role ORDER BY role")
    print("— dishes by role —")
    for r, c in cur.fetchall():
      print(f"{r}: {c}")

    cur.execute("SELECT COUNT(*) FROM dish_ingredients")
    print(f"— dish_ingredients rows — {cur.fetchone()[0]}")

  finally:
    conn.close()


if __name__ == "__main__":
  main()