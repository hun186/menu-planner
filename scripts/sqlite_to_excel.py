#cd /d D:\shared\TopicClassification\menu-planner
#python scripts\sqlite_to_excel.py data\menu.db data\menu_export.xlsx
import argparse
import sqlite3
from pathlib import Path
from typing import List, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def _safe_sheet_title(title: str) -> str:
    bad = ['\\', '/', '?', '*', '[', ']', ':']
    for ch in bad:
        title = title.replace(ch, '_')
    title = title.strip() or "sheet"
    return title[:31]


def _fetch_all(conn: sqlite3.Connection, sql: str) -> Tuple[List[str], List[tuple]]:
    cur = conn.execute(sql)
    cols = [d[0] for d in cur.description]
    rows = cur.fetchall()
    return cols, rows


def _write_sheet(wb: Workbook, sheet_name: str, cols: List[str], rows: List[tuple]) -> None:
    ws = wb.create_sheet(title=_safe_sheet_title(sheet_name))
    ws.append(cols)

    for r in rows:
        ws.append(list(r))

    ws.freeze_panes = "A2"

    if cols:
        last_col = get_column_letter(len(cols))
        last_row = max(1, len(rows) + 1)
        ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    sample_n = min(200, len(rows))
    for j, col in enumerate(cols, start=1):
        max_len = len(str(col))
        for i in range(sample_n):
            v = rows[i][j - 1]
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(j)].width = min(60, max(10, max_len + 2))


def export_db_to_excel(conn: sqlite3.Connection, xlsx_path: Path) -> None:
    wb = Workbook()

    # 移除預設空白 sheet
    default_ws = wb.active
    wb.remove(default_ws)

    table_queries = [
        ("meta", "SELECT key, value FROM meta ORDER BY key"),
        ("ingredients", "SELECT id, name, category, protein_group, default_unit FROM ingredients ORDER BY id"),
        ("dishes", "SELECT id, name, role, cuisine, meat_type, tags_json FROM dishes ORDER BY role, id"),
        ("dish_ingredients", "SELECT dish_id, ingredient_id, qty, unit FROM dish_ingredients ORDER BY dish_id, ingredient_id"),
        ("ingredient_prices", "SELECT ingredient_id, price_date, price_per_unit, unit FROM ingredient_prices ORDER BY price_date, ingredient_id"),
        ("inventory", "SELECT ingredient_id, qty_on_hand, unit, updated_at, expiry_date FROM inventory ORDER BY ingredient_id"),
        ("unit_conversions", "SELECT from_unit, to_unit, factor FROM unit_conversions ORDER BY from_unit, to_unit"),
    ]

    for sheet_name, sql in table_queries:
        cols, rows = _fetch_all(conn, sql)
        _write_sheet(wb, sheet_name, cols, rows)

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(xlsx_path))


def main():
    parser = argparse.ArgumentParser(description="把 SQLite 資料庫各資料表匯出成 Excel")
    parser.add_argument("db_path", type=str, help="SQLite DB 路徑，例如 data/menu.db")
    parser.add_argument("xlsx_path", type=str, help="輸出 Excel 路徑，例如 data/menu_export.xlsx")
    args = parser.parse_args()

    db_path = Path(args.db_path).resolve()
    xlsx_path = Path(args.xlsx_path).resolve()

    if not db_path.exists():
        raise FileNotFoundError(f"找不到 SQLite 檔案：{db_path}")

    conn = sqlite3.connect(str(db_path))
    try:
        export_db_to_excel(conn, xlsx_path)
    finally:
        conn.close()

    print(f"✅ 已輸出 Excel：{xlsx_path}")


if __name__ == "__main__":
    main()