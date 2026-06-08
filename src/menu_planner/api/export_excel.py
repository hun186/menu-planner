# src/menu_planner/api/export_excel.py
from __future__ import annotations

import io
import json
from datetime import datetime, date
from typing import Any, Dict, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from ..engine.roles import ROLE_LABELS, ROLE_ORDER, ROLE_PLURALS
from .export_excel_breakdown import build_human_breakdown
from .export_excel_sheets import (
    append_config_sheet,
    append_procurement_sheet,
    append_procurement_summary_sheet,
    append_summary_sheet,
    auto_fit_columns,
)


def _to_float_or_none(value: Any) -> float | None:
    try:
        if value is None or value == "":
            return None
        return float(value)
    except Exception:
        return None


ROLE_EXPORT_ORDER = ROLE_ORDER
METRIC_HEADERS = ["成本", "目標匹配度", "分數拆解(JSON)", "分數拆解(易讀)"]
WEEKDAY_SHORT_LABELS = {
    1: "一",
    2: "二",
    3: "三",
    4: "四",
    5: "五",
    6: "六",
    7: "日",
}
WEEKEND_OFFDAY_FILL = "FFFFE4E6"


def _iso_weekday_from_date_text(value: Any) -> int | None:
    if not isinstance(value, str):
        return None
    try:
        return date.fromisoformat(value).isoweekday()
    except ValueError:
        return None


def _weekday_short_label(value: Any) -> str:
    weekday = _iso_weekday_from_date_text(value)
    return WEEKDAY_SHORT_LABELS.get(weekday, "") if weekday is not None else ""


def _is_weekend_offday(day: Dict[str, Any]) -> bool:
    is_scheduled = day.get("is_scheduled")
    if is_scheduled is None:
        is_scheduled = True
    weekday = _iso_weekday_from_date_text(day.get("date"))
    return not bool(is_scheduled) and weekday in (6, 7)


def _as_nonnegative_int(value: Any) -> int:
    try:
        return max(0, int(value))
    except Exception:
        return 0


def _iter_role_count_maps(cfg: Dict[str, Any]) -> Iterable[Dict[str, Any]]:
    per_day = cfg.get("per_day_roles")
    if isinstance(per_day, dict):
        yield per_day

    per_weekday = cfg.get("per_weekday_roles")
    if isinstance(per_weekday, dict):
        for value in per_weekday.values():
            if isinstance(value, dict):
                yield value


def _role_dishes(items: Dict[str, Any], role: str) -> list[Dict[str, Any]]:
    plural = ROLE_PLURALS[role]
    dishes = items.get(plural)
    if isinstance(dishes, list):
        filtered = [x for x in dishes if isinstance(x, dict) and (x.get("name") or x.get("id"))]
        if filtered:
            return filtered

    dish = items.get(role) or {}
    if isinstance(dish, dict) and (dish.get("name") or dish.get("id")):
        return [dish]
    return []


def _compute_role_slots(cfg: Dict[str, Any], days: list[Dict[str, Any]]) -> Dict[str, int]:
    slots = {role: 0 for role in ROLE_EXPORT_ORDER}

    for counts in _iter_role_count_maps(cfg):
        for role in ROLE_EXPORT_ORDER:
            slots[role] = max(slots[role], _as_nonnegative_int(counts.get(role)))

    for day in days:
        items = day.get("items") or {}
        if not isinstance(items, dict):
            continue
        for role in ROLE_EXPORT_ORDER:
            slots[role] = max(slots[role], len(_role_dishes(items, role)))

    # Keep familiar single-slot roles visible even when all generated days are
    # off-days/failed, while allowing optional roles like noodle to disappear
    # when neither the config nor result asks for them.
    for role in ("main", "soup", "fruit"):
        slots[role] = max(slots[role], 1)

    return slots


def _role_headers(role_slots: Dict[str, int]) -> list[str]:
    headers: list[str] = []
    for role in ROLE_EXPORT_ORDER:
        count = role_slots.get(role, 0)
        label = ROLE_LABELS[role]
        if count == 1:
            headers.append(label)
        else:
            headers.extend(f"{label}{idx}" for idx in range(1, count + 1))
    return headers


def _extract_role_names(items: Dict[str, Any], role: str, count: int) -> list[str]:
    names = [(x.get("name") or x.get("id") or "") for x in _role_dishes(items, role)]
    if len(names) < count:
        names.extend([""] * (count - len(names)))
    return names[:count]


def _as_positive_int(value: Any, fallback: int) -> int:
    try:
        return max(1, int(float(value)))
    except Exception:
        return fallback


def _resolve_day_people(cfg: Dict[str, Any], day: Dict[str, Any], day_index: int) -> int:
    procurement = day.get("procurement") if isinstance(day, dict) else None
    if isinstance(procurement, dict) and procurement.get("people") not in (None, ""):
        return _as_positive_int(procurement.get("people"), 250)

    default_people = _as_positive_int((cfg or {}).get("people", 250), 250)
    schedule = (cfg or {}).get("schedule") or {}
    overrides = schedule.get("people_overrides") if isinstance(schedule, dict) else {}
    if isinstance(overrides, dict):
        date_key = str(day.get("date", ""))
        override = overrides.get(date_key)
        if override in (None, ""):
            override = overrides.get(str(day_index))
        if override not in (None, ""):
            return _as_positive_int(override, default_people)
    return default_people


def _extract_menu_row(
    cfg: Dict[str, Any],
    day: Dict[str, Any],
    role_slots: Dict[str, int],
    day_index: int,
) -> list[Any]:
    items = day.get("items", {}) or {}
    if not isinstance(items, dict):
        items = {}

    role_values: list[Any] = []
    for role in ROLE_EXPORT_ORDER:
        role_values.extend(_extract_role_names(items, role, role_slots.get(role, 0)))

    breakdown = day.get("score_breakdown") or {}
    breakdown_str = json.dumps(breakdown, ensure_ascii=False)

    fitness = day.get("score_fitness")
    if fitness is None and day.get("score") is not None:
        fitness = round(-float(day["score"]), 2)

    human = build_human_breakdown(day) if not day.get("failed") else ""
    return [
        day.get("date", ""),
        _weekday_short_label(day.get("date")),
        _resolve_day_people(cfg, day, day_index),
        *role_values,
        day.get("day_cost", ""),
        fitness if fitness is not None else "",
        breakdown_str,
        human,
    ]


def _compute_plan_totals(days: list[Dict[str, Any]]) -> tuple[float, float, int]:
    total_cost = 0.0
    total_fitness = 0.0
    fitness_count = 0

    for day in days:
        cost = _to_float_or_none(day.get("day_cost"))
        if cost is not None:
            total_cost += cost

        if day.get("failed"):
            continue

        fitness = _to_float_or_none(day.get("score_fitness"))
        if fitness is None:
            raw_score = _to_float_or_none(day.get("score"))
            if raw_score is not None:
                fitness = -raw_score

        if fitness is not None:
            total_fitness += fitness
            fitness_count += 1

    return total_cost, total_fitness, fitness_count


def build_plan_workbook(cfg: Dict[str, Any], result: Dict[str, Any]) -> bytes:
    """
    result 格式：engine/explain.py build_explanations 的輸出
    """
    wb = Workbook()

    ws = wb.active
    ws.title = "菜單"
    days = result.get("days", []) or []
    role_slots = _compute_role_slots(cfg, days)
    header = ["日期", "週幾", "人數", *_role_headers(role_slots), *METRIC_HEADERS]
    ws.append(header)

    bold = Font(bold=True)
    for col in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = bold
        cell.alignment = Alignment(vertical="center")

    weekend_offday_fill = PatternFill(fill_type="solid", fgColor=WEEKEND_OFFDAY_FILL)
    for day_index, day in enumerate(days):
        ws.append(_extract_menu_row(cfg, day, role_slots, day_index))
        if _is_weekend_offday(day):
            for cell in ws[ws.max_row]:
                cell.fill = weekend_offday_fill

    total_cost, total_fitness, fitness_count = _compute_plan_totals(days)

    ws.freeze_panes = "A2"
    auto_fit_columns(ws, max_width=80)

    wrap = Alignment(vertical="top", wrap_text=True)
    human_col = len(header)
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=human_col).alignment = wrap

    append_summary_sheet(wb, cfg, result, days, total_cost, total_fitness, fitness_count)
    append_config_sheet(wb, cfg)
    append_procurement_sheet(wb, result)
    append_procurement_summary_sheet(wb, result)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()



def build_filename(prefix: str = "menu_plan") -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{ts}.xlsx"
