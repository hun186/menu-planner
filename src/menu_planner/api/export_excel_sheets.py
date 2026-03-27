from __future__ import annotations

import json
from typing import Any, Dict, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .export_excel_breakdown import num

DAILY_SUBTOTAL_FILL = PatternFill(fill_type="solid", fgColor="FFFDE68A")
WEEKLY_SUBTOTAL_FILL = PatternFill(fill_type="solid", fgColor="FFBFDBFE")
DAILY_SUBTOTAL_FONT = Font(color="FF92400E", bold=True)
WEEKLY_SUBTOTAL_FONT = Font(color="FF1E3A8A", bold=True)


def set_col_width(ws, widths: Dict[int, float]) -> None:
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def append_procurement_sheet(wb: Workbook, result: Dict[str, Any]) -> None:
    ws = wb.create_sheet("採買明細")
    header = [
        "日期", "角色", "菜名", "食材", "每人用量", "人數", "需求量", "需求單位",
        "單價", "單價單位", "小計", "價格日期",
    ]
    ws.append(header)

    bold = Font(bold=True)
    for col in range(1, len(header) + 1):
        ws.cell(row=1, column=col).font = bold

    for day in (result.get("days") or []):
        procurement = day.get("procurement") or {}
        date_text = day.get("date", "")
        people = procurement.get("people", 250)
        for dish in (procurement.get("dishes") or []):
            role = dish.get("role", "")
            dish_name = dish.get("dish_name", "")
            for ingredient in (dish.get("ingredients") or []):
                ws.append([
                    date_text,
                    role,
                    dish_name,
                    ingredient.get("ingredient_name", ""),
                    ingredient.get("qty_per_person", ""),
                    people,
                    ingredient.get("qty_for_people", ""),
                    ingredient.get("qty_unit", ""),
                    ingredient.get("unit_price", ""),
                    ingredient.get("unit_price_unit", ""),
                    ingredient.get("line_total", ""),
                    ingredient.get("price_date", ""),
                ])

    ws.freeze_panes = "A2"
    set_col_width(ws, {1: 12, 2: 10, 3: 20, 4: 18, 5: 12, 6: 8, 7: 12, 8: 10, 9: 10, 10: 12, 11: 10, 12: 12})


def append_procurement_summary_sheet(wb: Workbook, result: Dict[str, Any]) -> None:
    ws = wb.create_sheet("採買彙總")
    header = ["週次", "日期", "食材", "單價", "單價單位", "總量", "需求單位", "總價格", "備註"]
    ws.append(header)

    bold = Font(bold=True)
    for col in range(1, len(header) + 1):
        ws.cell(row=1, column=col).font = bold

    grand_total = 0.0
    week_total = 0.0
    week_index = 1
    valid_days = [day for day in (result.get("days") or []) if day.get("procurement")]

    for day_idx, day in enumerate(valid_days):
        procurement = day.get("procurement") or {}
        date_text = day.get("date", "")
        people = procurement.get("people", "")
        agg: Dict[Tuple[str, str, str, float], Dict[str, Any]] = {}

        for dish in (procurement.get("dishes") or []):
            for ingredient in (dish.get("ingredients") or []):
                unit_price = num(ingredient.get("unit_price"), 0.0)
                key = (
                    str(ingredient.get("ingredient_name", "")),
                    str(ingredient.get("qty_unit", "")),
                    str(ingredient.get("unit_price_unit", "")),
                    unit_price,
                )
                bucket = agg.setdefault(key, {"qty": 0.0, "total": 0.0})
                bucket["qty"] += num(ingredient.get("qty_for_people"), 0.0)
                bucket["total"] += num(ingredient.get("line_total"), 0.0)

        day_total = 0.0
        for (name, qty_unit, price_unit, unit_price), value in sorted(agg.items(), key=lambda kv: kv[0][0]):
            total = round(value["total"], 2)
            day_total += total
            ws.append([
                f"第{week_index}週",
                date_text,
                name,
                round(unit_price, 4) if unit_price else "",
                price_unit,
                round(value["qty"], 4),
                qty_unit,
                total,
                f"人數={people}",
            ])

        ws.append([f"第{week_index}週", date_text, "每日小計", "", "", "", "", round(day_total, 2), ""])
        daily_row = ws.max_row
        for col in range(1, len(header) + 1):
            cell = ws.cell(row=daily_row, column=col)
            cell.fill = DAILY_SUBTOTAL_FILL
            cell.font = DAILY_SUBTOTAL_FONT
        week_total += day_total
        grand_total += day_total

        week_done = ((day_idx + 1) % 7 == 0) or (day_idx == len(valid_days) - 1)
        if week_done:
            ws.append([f"第{week_index}週", "", "每週小計", "", "", "", "", round(week_total, 2), ""])
            weekly_row = ws.max_row
            for col in range(1, len(header) + 1):
                cell = ws.cell(row=weekly_row, column=col)
                cell.fill = WEEKLY_SUBTOTAL_FILL
                cell.font = WEEKLY_SUBTOTAL_FONT
            week_index += 1
            week_total = 0.0

    ws.append(["", "", "全部合計", "", "", "", "", round(grand_total, 2), ""])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{ws.max_row}"
    set_col_width(ws, {1: 10, 2: 12, 3: 18, 4: 10, 5: 10, 6: 12, 7: 10, 8: 12, 9: 14})


def append_summary_sheet(
    wb: Workbook,
    cfg: Dict[str, Any],
    result: Dict[str, Any],
    days: list[Dict[str, Any]],
    total_cost: float,
    total_fitness: float,
    fitness_count: int,
) -> None:
    ws = wb.create_sheet("摘要")
    summary = result.get("summary", {}) or {}
    days_n = int(summary.get("days") or len(days) or 0)

    bold = Font(bold=True)
    ws.append(["項目", "值"])
    ws["A1"].font = bold
    ws["B1"].font = bold
    ws.append(["天數", days_n])
    ws.append(["人數", int((cfg or {}).get("people", 250) or 250)])
    ws.append(["總成本", round(total_cost, 2)])
    ws.append(["平均/日", round(total_cost / max(days_n, 1), 2)])
    ws.append(["總符合度", round(total_fitness, 2)])
    ws.append(["平均符合度/日", round(total_fitness / max(fitness_count, 1), 2)])
    ws.append(["說明", "符合度越高越好；若未提供 score_fitness，則使用 -原始分數 當符合度。"])

    set_col_width(ws, {1: 18, 2: 60})


def append_config_sheet(wb: Workbook, cfg: Dict[str, Any]) -> None:
    ws = wb.create_sheet("設定")
    bold = Font(bold=True)
    ws.append(["constraints.json"])
    ws["A1"].font = bold

    cfg_str = json.dumps(cfg, ensure_ascii=False, indent=2)
    for line in cfg_str.splitlines():
        ws.append([line])

    set_col_width(ws, {1: 110})
