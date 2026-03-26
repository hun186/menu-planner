#src/menu_planner/api/routes/admin_catalog.py
from __future__ import annotations

import os
import sqlite3
import shutil
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import List, Optional

from fastapi import APIRouter, Body, Depends, Header, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field

from ...db.admin_repo import SQLiteAdminRepo
from ...db.backup import (
    BACKUP_REASON_DEFAULT,
    create_db_backup,
    get_backup_metadata_map,
    remove_backup_metadata,
    upsert_backup_metadata,
)

DEFAULT_DB_PATH = str((__import__("pathlib").Path.cwd() / "data" / "menu.db").resolve())

router = APIRouter(prefix="/admin/catalog", tags=["admin-catalog"])
BACKUP_WARNING_THRESHOLD_BYTES = 500 * 1024 * 1024


def _timestamp_for_filename() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _build_excel_response(filename_prefix: str, sheet_name: str, headers: List[str], rows: List[List[object]]) -> StreamingResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] if sheet_name else "sheet1"
    ws.append(headers)
    for row in rows:
        ws.append(row)

    for idx, title in enumerate(headers, start=1):
        col = ws.column_dimensions[get_column_letter(idx)]
        col.width = max(12, min(40, len(str(title)) + 6))
        ws.cell(row=1, column=idx).font = Font(bold=True)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"{filename_prefix}_{_timestamp_for_filename()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def backup_before_modify(
    db_path: str,
    reason: str = BACKUP_REASON_DEFAULT,
    comment: str = "",
) -> None:
    try:
        create_db_backup(db_path, reason=reason, comment=comment)
    except FileNotFoundError:
        raise HTTPException(status_code=500, detail=f"資料庫檔案不存在：{db_path}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"建立資料庫備份失敗：{e}")


def repo_with_backup(
    db_path: str,
    *,
    reason: str = BACKUP_REASON_DEFAULT,
    comment: str = "",
) -> SQLiteAdminRepo:
    backup_before_modify(db_path, reason=reason, comment=comment)
    return SQLiteAdminRepo(db_path)


def ensure_ingredient_exists(repo: SQLiteAdminRepo, ingredient_id: str) -> None:
    if not repo.ingredient_exists(ingredient_id):
        raise HTTPException(status_code=404, detail="找不到此食材")


def ensure_dish_exists(repo: SQLiteAdminRepo, dish_id: str) -> None:
    if not repo.dish_exists(dish_id):
        raise HTTPException(status_code=404, detail="找不到此菜色")


def require_admin_key(x_admin_key: Optional[str] = Header(default=None)):
    required = os.getenv("MENU_ADMIN_KEY")
    if required and x_admin_key != required:
        raise HTTPException(status_code=401, detail="未授權：需要有效的 X-Admin-Key")


class IngredientUpsert(BaseModel):
    name: str = Field(min_length=1)
    category: str = Field(min_length=1)
    protein_group: Optional[str] = None
    default_unit: str = Field(min_length=1)


class DishUpsert(BaseModel):
    name: str = Field(min_length=1)
    role: str = Field(pattern="^(main|side|veg|soup|fruit)$")
    cuisine: Optional[str] = None
    meat_type: Optional[str] = None
    tags: List[str] = Field(default_factory=list)


class DishIngredientIn(BaseModel):
    ingredient_id: str = Field(min_length=1)
    qty: float = Field(gt=0)
    unit: str = Field(min_length=1)


class DishCostPreviewIn(BaseModel):
    items: List[DishIngredientIn] = Field(default_factory=list)
    servings: float = Field(default=1.0, gt=0)


class BackupRestoreIn(BaseModel):
    backup_filename: str = Field(min_length=1)


class BackupCommentIn(BaseModel):
    comment: str = Field(default="", max_length=500)


class BackupCreateIn(BaseModel):
    reason: str = Field(default="admin_manual_snapshot", max_length=120)
    comment: str = Field(default="", max_length=500)


class DishRenameIn(DishUpsert):
    target_id: str = Field(min_length=1)


@router.get("/ingredients", dependencies=[Depends(require_admin_key)])
def list_ingredients(
    q: Optional[str] = Query(default=None),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=200),
    db_path: str = Query(default=DEFAULT_DB_PATH),
):
    repo = SQLiteAdminRepo(db_path)
    return repo.list_ingredients(q=q, page=page, page_size=page_size)


@router.get("/dishes", dependencies=[Depends(require_admin_key)])
def list_dishes(
    q: Optional[str] = Query(default=None),
    role: Optional[str] = Query(default=None),
    ingredient_id: Optional[str] = Query(default=None),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=200),
    db_path: str = Query(default=DEFAULT_DB_PATH),
):
    repo = SQLiteAdminRepo(db_path)
    return repo.list_dishes(q=q, role=role, ingredient_id=ingredient_id, page=page, page_size=page_size)


@router.put("/ingredients/{ingredient_id}", dependencies=[Depends(require_admin_key)])
def upsert_ingredient(
    ingredient_id: str,
    body: IngredientUpsert,
    db_path: str = Query(default=DEFAULT_DB_PATH),
):
    repo = repo_with_backup(db_path, reason="ingredient_upsert")
    repo.upsert_ingredient(ingredient_id, body.model_dump())
    return {"ok": True, "id": ingredient_id}


@router.delete("/ingredients/{ingredient_id}", dependencies=[Depends(require_admin_key)])
def delete_ingredient(
    ingredient_id: str,
    db_path: str = Query(default=DEFAULT_DB_PATH),
):
    repo = SQLiteAdminRepo(db_path)
    ensure_ingredient_exists(repo, ingredient_id)

    repo = repo_with_backup(db_path, reason="ingredient_delete")
    try:
        n = repo.delete_ingredient(ingredient_id)
        if n == 0:
            raise HTTPException(status_code=404, detail="找不到此食材")
        return {"ok": True}
    except sqlite3.IntegrityError:
        refs = repo.find_dishes_using_ingredient(ingredient_id)
        raise HTTPException(
            status_code=409,
            detail={"message": "此食材已被菜色引用，無法刪除", "referenced_by": refs},
        )


@router.post("/ingredients/{ingredient_id}/rename", dependencies=[Depends(require_admin_key)])
def rename_ingredient(
    ingredient_id: str,
    body: IngredientRenameIn,
    db_path: str = Query(default=DEFAULT_DB_PATH),
):
    repo = SQLiteAdminRepo(db_path)
    ensure_ingredient_exists(repo, ingredient_id)
    target_id = str(body.target_id or "").strip()
    if not target_id:
        raise HTTPException(status_code=400, detail="target_id 不可為空")
    if target_id == ingredient_id:
        raise HTTPException(status_code=400, detail="target_id 不可與來源 ingredient_id 相同")

    repo = repo_with_backup(db_path, reason="ingredient_rename")
    try:
        result = repo.rename_ingredient(
            ingredient_id,
            target_id,
            IngredientUpsert(
                name=body.name,
                category=body.category,
                protein_group=body.protein_group,
                default_unit=body.default_unit,
            ).model_dump(),
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"ok": True, **result}


class PriceUpsert(BaseModel):
    price_per_unit: float = Field(gt=0)
    unit: str = Field(min_length=1)


class InventoryUpsert(BaseModel):
    qty_on_hand: float = Field(ge=0)
    unit: str = Field(min_length=1)
    updated_at: str = Field(min_length=10)  # YYYY-MM-DD
    expiry_date: Optional[str] = None  # YYYY-MM-DD or null


class IngredientMergeIn(BaseModel):
    source_ingredient_id: str = Field(min_length=1)
    target_ingredient_id: str = Field(min_length=1)


class IngredientRenameIn(IngredientUpsert):
    target_id: str = Field(min_length=1)


@router.get("/ingredients/{ingredient_id}/prices", dependencies=[Depends(require_admin_key)])
def list_prices(
    ingredient_id: str,
    limit: int = Query(default=30, ge=1, le=365),
    db_path: str = Query(default=DEFAULT_DB_PATH),
):
    repo = SQLiteAdminRepo(db_path)
    ensure_ingredient_exists(repo, ingredient_id)
    return repo.list_prices(ingredient_id, limit=limit)


@router.put("/ingredients/{ingredient_id}/prices/{price_date}", dependencies=[Depends(require_admin_key)])
def upsert_price(
    ingredient_id: str,
    price_date: str,
    body: PriceUpsert,
    db_path: str = Query(default=DEFAULT_DB_PATH),
):
    repo = SQLiteAdminRepo(db_path)
    ensure_ingredient_exists(repo, ingredient_id)

    repo = repo_with_backup(db_path, reason="ingredient_price_upsert")
    repo.upsert_price(ingredient_id, price_date, body.model_dump())
    return {"ok": True}


@router.delete("/ingredients/{ingredient_id}/prices/{price_date}", dependencies=[Depends(require_admin_key)])
def delete_price(
    ingredient_id: str,
    price_date: str,
    db_path: str = Query(default=DEFAULT_DB_PATH),
):
    repo = SQLiteAdminRepo(db_path)
    ensure_ingredient_exists(repo, ingredient_id)
    if not repo.price_exists(ingredient_id, price_date):
        raise HTTPException(status_code=404, detail="找不到此價格紀錄")

    repo = repo_with_backup(db_path, reason="ingredient_price_delete")
    repo.delete_price(ingredient_id, price_date)
    return {"ok": True}


@router.get("/ingredients/{ingredient_id}/inventory", dependencies=[Depends(require_admin_key)])
def get_inventory(
    ingredient_id: str,
    db_path: str = Query(default=DEFAULT_DB_PATH),
):
    repo = SQLiteAdminRepo(db_path)
    ensure_ingredient_exists(repo, ingredient_id)
    return repo.get_inventory(ingredient_id)  # 可能回 null


@router.put("/ingredients/{ingredient_id}/inventory", dependencies=[Depends(require_admin_key)])
def upsert_inventory(
    ingredient_id: str,
    body: InventoryUpsert,
    db_path: str = Query(default=DEFAULT_DB_PATH),
):
    repo = SQLiteAdminRepo(db_path)
    ensure_ingredient_exists(repo, ingredient_id)

    repo = repo_with_backup(db_path, reason="ingredient_inventory_upsert")
    repo.upsert_inventory(ingredient_id, body.model_dump())
    return {"ok": True}


@router.get("/inventory/summary", dependencies=[Depends(require_admin_key)])
def list_inventory_summary(
    q: Optional[str] = Query(default=None),
    only_in_stock: bool = Query(default=False),
    db_path: str = Query(default=DEFAULT_DB_PATH),
):
    repo = SQLiteAdminRepo(db_path)
    return repo.list_inventory_summary(q=q, only_in_stock=only_in_stock)


def _list_backup_files(db_path: str) -> List[dict]:
    db_file = Path(db_path).resolve()
    backup_dir = db_file.parent / "backups"
    if not backup_dir.exists():
        return []
    pattern = f"{db_file.stem}_*{db_file.suffix or '.db'}"
    files = sorted(backup_dir.glob(pattern), key=lambda p: p.name, reverse=True)
    metadata_map = get_backup_metadata_map(db_path)
    return [
        {
            "filename": p.name,
            "size_bytes": p.stat().st_size,
            "modified_at": datetime.fromtimestamp(p.stat().st_mtime).isoformat(timespec="seconds"),
            "action_reason": str((metadata_map.get(p.name) or {}).get("reason") or BACKUP_REASON_DEFAULT),
            "comment": str((metadata_map.get(p.name) or {}).get("comment") or ""),
        }
        for p in files
        if p.is_file()
    ]


def _summarize_backup_usage(files: List[dict]) -> dict:
    total_bytes = sum(int(x.get("size_bytes") or 0) for x in files)
    return {
        "count": len(files),
        "total_size_bytes": total_bytes,
        "warning_threshold_bytes": BACKUP_WARNING_THRESHOLD_BYTES,
        "is_over_warning_threshold": total_bytes >= BACKUP_WARNING_THRESHOLD_BYTES,
    }


@router.get("/backups", dependencies=[Depends(require_admin_key)])
def list_db_backups(
    db_path: str = Query(default=DEFAULT_DB_PATH),
):
    return _list_backup_files(db_path)


@router.get("/backups/stats", dependencies=[Depends(require_admin_key)])
def get_db_backup_stats(
    db_path: str = Query(default=DEFAULT_DB_PATH),
):
    files = _list_backup_files(db_path)
    return _summarize_backup_usage(files)


@router.post("/backups/create", dependencies=[Depends(require_admin_key)])
def create_manual_db_backup(
    body: BackupCreateIn,
    db_path: str = Query(default=DEFAULT_DB_PATH),
):
    reason = str(body.reason or "admin_manual_snapshot").strip() or "admin_manual_snapshot"
    comment = str(body.comment or "").strip()
    backup_before_modify(db_path, reason=reason, comment=comment)
    return {"ok": True, "reason": reason, "comment": comment}


@router.post("/backups/restore", dependencies=[Depends(require_admin_key)])
def restore_db_backup(
    body: BackupRestoreIn,
    db_path: str = Query(default=DEFAULT_DB_PATH),
):
    db_file = Path(db_path).resolve()
    backup_dir = db_file.parent / "backups"
    backup_name = body.backup_filename.strip()
    if not backup_name or "/" in backup_name or "\\" in backup_name:
        raise HTTPException(status_code=400, detail="備份檔名格式不正確")
    if not backup_name.startswith(f"{db_file.stem}_") or not backup_name.endswith(db_file.suffix or ".db"):
        raise HTTPException(status_code=400, detail="備份檔名不符合目前資料庫")

    src = (backup_dir / backup_name).resolve()
    if src.parent != backup_dir.resolve() or not src.exists() or not src.is_file():
        raise HTTPException(status_code=404, detail="找不到指定備份檔")

    backup_before_modify(str(db_file), reason="admin_restore_pre_snapshot")
    try:
        shutil.copy2(src, db_file)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"還原備份失敗：{e}")
    return {"ok": True, "restored_from": backup_name}


@router.delete("/backups/{backup_name}", dependencies=[Depends(require_admin_key)])
def delete_db_backup(
    backup_name: str,
    db_path: str = Query(default=DEFAULT_DB_PATH),
):
    db_file = Path(db_path).resolve()
    backup_dir = db_file.parent / "backups"
    name = backup_name.strip()
    if not name or "/" in name or "\\" in name:
        raise HTTPException(status_code=400, detail="備份檔名格式不正確")
    if not name.startswith(f"{db_file.stem}_") or not name.endswith(db_file.suffix or ".db"):
        raise HTTPException(status_code=400, detail="備份檔名不符合目前資料庫")

    target = (backup_dir / name).resolve()
    if target.parent != backup_dir.resolve() or not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail="找不到指定備份檔")

    try:
        target.unlink()
        remove_backup_metadata(db_path, name)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"刪除備份失敗：{e}")
    return {"ok": True, "deleted": name}


@router.patch("/backups/{backup_name}/comment", dependencies=[Depends(require_admin_key)])
def update_db_backup_comment(
    backup_name: str,
    body: BackupCommentIn,
    db_path: str = Query(default=DEFAULT_DB_PATH),
):
    db_file = Path(db_path).resolve()
    backup_dir = db_file.parent / "backups"
    name = backup_name.strip()
    if not name or "/" in name or "\\" in name:
        raise HTTPException(status_code=400, detail="備份檔名格式不正確")
    if not name.startswith(f"{db_file.stem}_") or not name.endswith(db_file.suffix or ".db"):
        raise HTTPException(status_code=400, detail="備份檔名不符合目前資料庫")
    target = (backup_dir / name).resolve()
    if target.parent != backup_dir.resolve() or not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail="找不到指定備份檔")

    existing_meta = get_backup_metadata_map(db_path).get(name) or {}
    upsert_backup_metadata(
        db_path=db_path,
        backup_filename=name,
        reason=str(existing_meta.get("reason") or BACKUP_REASON_DEFAULT),
        comment=body.comment,
    )
    return {"ok": True, "filename": name, "comment": body.comment}


@router.post("/inventory/summary/merge-ingredient", dependencies=[Depends(require_admin_key)])
def merge_inventory_ingredient(
    body: IngredientMergeIn,
    db_path: str = Query(default=DEFAULT_DB_PATH),
):
    source_id = body.source_ingredient_id.strip()
    target_id = body.target_ingredient_id.strip()
    if source_id == target_id:
        raise HTTPException(status_code=400, detail="來源與目標食材不可相同")

    repo = SQLiteAdminRepo(db_path)
    ensure_ingredient_exists(repo, source_id)
    ensure_ingredient_exists(repo, target_id)

    repo = repo_with_backup(
        db_path,
        reason=f"ingredient_merge:{source_id}->{target_id}",
    )
    try:
        result = repo.merge_ingredient(source_id, target_id)
        return {"ok": True, **result}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/inventory/summary/export", dependencies=[Depends(require_admin_key)])
def export_inventory_summary_excel(
    q: Optional[str] = Query(default=None),
    only_in_stock: bool = Query(default=False),
    db_path: str = Query(default=DEFAULT_DB_PATH),
):
    repo = SQLiteAdminRepo(db_path)
    rows_raw = repo.list_inventory_summary(q=q, only_in_stock=only_in_stock)
    rows = [
        [
            r.get("ingredient_id"),
            r.get("ingredient_name"),
            r.get("category"),
            r.get("qty_on_hand"),
            r.get("inventory_unit") or r.get("default_unit"),
            r.get("updated_at"),
            r.get("expiry_date"),
            r.get("dish_ref_count"),
        ]
        for r in rows_raw
    ]
    return _build_excel_response(
        filename_prefix="inventory_summary",
        sheet_name="庫存統整",
        headers=["食材ID", "名稱", "分類", "庫存量", "庫存單位", "更新日", "到期日", "引用菜色數"],
        rows=rows,
    )


@router.get("/ingredients/export", dependencies=[Depends(require_admin_key)])
def export_ingredients_excel(
    q: Optional[str] = Query(default=None),
    db_path: str = Query(default=DEFAULT_DB_PATH),
):
    repo = SQLiteAdminRepo(db_path)
    payload = repo.list_ingredients(q=q, page=1, page_size=100000)
    items = payload.get("items") or []
    rows = [
        [r.get("id"), r.get("name"), r.get("category"), r.get("protein_group"), r.get("default_unit")]
        for r in items
    ]
    return _build_excel_response(
        filename_prefix="ingredients",
        sheet_name="食材管理",
        headers=["食材ID", "名稱", "分類", "蛋白群組", "預設單位"],
        rows=rows,
    )


@router.get("/dishes/export", dependencies=[Depends(require_admin_key)])
def export_dishes_excel(
    q: Optional[str] = Query(default=None),
    ingredient_id: Optional[str] = Query(default=None),
    role: Optional[str] = Query(default=None),
    db_path: str = Query(default=DEFAULT_DB_PATH),
):
    repo = SQLiteAdminRepo(db_path)
    payload = repo.list_dishes(q=q, role=role, ingredient_id=ingredient_id, page=1, page_size=100000)
    items = payload.get("items") or []
    rows = [
        [
            r.get("id"),
            r.get("name"),
            r.get("role"),
            r.get("meat_type"),
            r.get("cuisine"),
            ",".join(r.get("tags") or []),
        ]
        for r in items
    ]
    return _build_excel_response(
        filename_prefix="dishes",
        sheet_name="菜名管理",
        headers=["菜色ID", "名稱", "角色", "肉類", "菜系", "標籤"],
        rows=rows,
    )


@router.put("/dishes/{dish_id}", dependencies=[Depends(require_admin_key)])
def upsert_dish(
    dish_id: str,
    body: DishUpsert,
    db_path: str = Query(default=DEFAULT_DB_PATH),
):
    repo = repo_with_backup(db_path, reason="dish_upsert")
    repo.upsert_dish(dish_id, body.model_dump())
    return {"ok": True, "id": dish_id}


@router.delete("/dishes/{dish_id}", dependencies=[Depends(require_admin_key)])
def delete_dish(
    dish_id: str,
    db_path: str = Query(default=DEFAULT_DB_PATH),
):
    repo = SQLiteAdminRepo(db_path)
    ensure_dish_exists(repo, dish_id)

    repo = repo_with_backup(db_path, reason="dish_delete")
    n = repo.delete_dish(dish_id)
    if n == 0:
        raise HTTPException(status_code=404, detail="找不到此菜色")
    return {"ok": True}


@router.post("/dishes/{dish_id}/rename", dependencies=[Depends(require_admin_key)])
def rename_dish(
    dish_id: str,
    body: DishRenameIn,
    db_path: str = Query(default=DEFAULT_DB_PATH),
):
    repo = SQLiteAdminRepo(db_path)
    ensure_dish_exists(repo, dish_id)
    target_id = str(body.target_id or "").strip()
    if not target_id:
        raise HTTPException(status_code=400, detail="target_id 不可為空")
    if target_id == dish_id:
        raise HTTPException(status_code=400, detail="target_id 不可與來源 dish_id 相同")

    repo = repo_with_backup(db_path, reason="dish_rename")
    try:
        result = repo.rename_dish(
            dish_id,
            target_id,
            DishUpsert(
                name=body.name,
                role=body.role,
                cuisine=body.cuisine,
                meat_type=body.meat_type,
                tags=body.tags,
            ).model_dump(),
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"ok": True, **result}


@router.get("/dishes/{dish_id}/ingredients", dependencies=[Depends(require_admin_key)])
def get_dish_ingredients(
    dish_id: str,
    db_path: str = Query(default=DEFAULT_DB_PATH),
):
    repo = SQLiteAdminRepo(db_path)
    ensure_dish_exists(repo, dish_id)
    return repo.get_dish_ingredients(dish_id)


@router.put("/dishes/{dish_id}/ingredients", dependencies=[Depends(require_admin_key)])
def put_dish_ingredients(
    dish_id: str,
    items: List[DishIngredientIn] = Body(...),
    db_path: str = Query(default=DEFAULT_DB_PATH),
):
    repo = SQLiteAdminRepo(db_path)
    ensure_dish_exists(repo, dish_id)

    missing = repo.find_missing_ingredients([x.ingredient_id for x in items])
    if missing:
        raise HTTPException(status_code=400, detail={"message": "有不存在的食材 id", "missing": missing})

    repo = repo_with_backup(db_path, reason="dish_ingredients_replace")
    repo.replace_dish_ingredients(dish_id, [x.model_dump() for x in items])
    return {"ok": True}


@router.post("/dishes/cost-preview", dependencies=[Depends(require_admin_key)])
def dish_cost_preview(
    body: DishCostPreviewIn,
    db_path: str = Query(default=DEFAULT_DB_PATH),
):
    repo = SQLiteAdminRepo(db_path)
    return repo.preview_dish_cost([x.model_dump() for x in body.items], servings=body.servings)


@router.get("/dishes/cost-preview", dependencies=[Depends(require_admin_key)])
def list_dish_cost_preview(
    dish_id: List[str] = Query(default=[]),
    db_path: str = Query(default=DEFAULT_DB_PATH),
):
    repo = SQLiteAdminRepo(db_path)
    dish_ids = dish_id if isinstance(dish_id, list) else []
    target_dish_ids = dish_ids or None
    if target_dish_ids is None:
        return repo.list_dish_cost_preview()
    return repo.list_dish_cost_preview(target_dish_ids)
