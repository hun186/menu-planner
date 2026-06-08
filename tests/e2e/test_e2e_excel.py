import io
import json
import urllib.request
from collections import Counter
from datetime import date, datetime, timedelta

import openpyxl

BASE = "http://127.0.0.1:18000"
ROLE_PLURALS = {
    "main": "mains",
    "noodle": "noodles",
    "side": "sides",
    "veg": "vegs",
    "soup": "soups",
    "fruit": "fruits",
}
ROLE_HEADER_LABELS = {
    "main": "主菜",
    "noodle": "麵食",
    "side": "配菜",
    "veg": "純蔬",
    "soup": "湯",
    "fruit": "水果",
}
ROLE_EXPORT_ORDER = ("main", "noodle", "side", "veg", "soup", "fruit")


def req(method, path, payload=None):
    data = None
    headers = {}

    if payload is not None:
        data = json.dumps(payload).encode()
        headers["Content-Type"] = "application/json"

    r = urllib.request.Request(BASE + path, data=data, headers=headers, method=method)

    with urllib.request.urlopen(r, timeout=180) as f:
        return f.status, dict(f.headers), f.read()


def _parse_start_date(cfg):
    raw = cfg.get("start_date")
    if raw:
        return datetime.strptime(raw, "%Y-%m-%d").date()
    return date.today()


def _role_counts_for_day(cfg, day_index):
    base = dict(cfg.get("per_day_roles") or {})
    current = _parse_start_date(cfg) + timedelta(days=day_index)
    override = (cfg.get("per_weekday_roles") or {}).get(str(current.isoweekday()))
    if isinstance(override, dict):
        base.update(override)
    return {role: max(0, int(base.get(role, 0) or 0)) for role in ROLE_EXPORT_ORDER}


def _role_items(items, role):
    values = items.get(ROLE_PLURALS[role])
    if isinstance(values, list):
        return values
    one = items.get(role) or {}
    return [one] if one.get("id") else []


def _expected_role_headers(cfg):
    max_counts = {role: 0 for role in ROLE_EXPORT_ORDER}
    for role, count in (cfg.get("per_day_roles") or {}).items():
        if role in max_counts:
            max_counts[role] = max(max_counts[role], int(count or 0))
    for override in (cfg.get("per_weekday_roles") or {}).values():
        if isinstance(override, dict):
            for role, count in override.items():
                if role in max_counts:
                    max_counts[role] = max(max_counts[role], int(count or 0))
    for role in ("main", "soup", "fruit"):
        max_counts[role] = max(max_counts[role], 1)

    headers = []
    for role in ROLE_EXPORT_ORDER:
        count = max_counts[role]
        label = ROLE_HEADER_LABELS[role]
        if count == 1:
            headers.append(label)
        else:
            headers.extend(f"{label}{idx}" for idx in range(1, count + 1))
    return headers


def test_generate_and_export_excel():
    # 取得預設設定
    st, _, body = req("GET", "/config/default")
    assert st == 200

    cfg = json.loads(body)
    cfg["horizon_days"] = 270
    # 使用固定 seed + 關閉 local search，避免隨機鄰域替換造成 CI 偶發失敗
    cfg["seed"] = 7
    (cfg.setdefault("search", {}).setdefault("local_search", {}))["enabled"] = False
    # 本測試聚焦在 Excel 匯出與彈性角色欄位；放寬 veg 重複限制避免測試資料量影響匯出驗證。
    (cfg.setdefault("hard", {}).setdefault("repeat_limits", {}))["max_same_veg_in_7_days"] = 99

    # 生成菜單
    st, _, body = req("POST", "/plan", cfg)
    assert st == 200

    obj = json.loads(body)

    assert obj["ok"] is True
    assert len(obj["result"]["days"]) == 270
    assert obj["result"]["summary"].get("people") == cfg.get("people", 1)

    # 驗證每日組成會依 per_day_roles / per_weekday_roles 的彈性角色數調整。
    # 只檢查有排程且成功的日子（offday / failed 另有邏輯）
    days = obj["result"]["days"]
    for d in days:
        if d.get("failed"):
            continue
        items = d.get("items") or {}
        main = items.get("main") or {}
        if not main.get("id"):
            # offday
            continue


        procurement = d.get("procurement") or {}
        expected_people = (cfg.get("schedule", {}).get("people_overrides", {}) or {}).get(
            d.get("date"),
            cfg.get("people", 1),
        )
        assert procurement.get("people") == expected_people
        expected_counts = _role_counts_for_day(cfg, d.get("day_index", 0))
        for role in ROLE_EXPORT_ORDER:
            role_items = _role_items(items, role)
            assert len(role_items) == expected_counts[role]
            assert all((item or {}).get("id") for item in role_items)

    # 驗證 veg 重複規則（最近 7 個有排餐日）
    rep = (cfg.get("hard") or {}).get("repeat_limits") or {}
    max_veg_7 = int(rep.get("max_same_veg_in_7_days", rep.get("max_same_side_in_7_days", 1)))
    active_completed: list[dict] = []
    for d in days:
        if d.get("failed"):
            continue
        items = d.get("items") or {}
        main_id = ((items.get("main") or {}).get("id") or "").strip()
        veg_id = ((items.get("veg") or {}).get("id") or "").strip()
        if not main_id:
            # offday
            continue
        assert veg_id

        recent_veg_ids = [
            ((x.get("items") or {}).get("veg") or {}).get("id")
            for x in active_completed[-7:]
            if ((x.get("items") or {}).get("veg") or {}).get("id")
        ]
        c = Counter(recent_veg_ids)
        assert c.get(veg_id, 0) + 1 <= max_veg_7
        active_completed.append(d)

    # 匯出 Excel
    st, _, xb = req(
        "POST",
        "/export/excel",
        {
            "cfg": cfg,
            "result": obj["result"],
        },
    )

    assert st == 200

    # 檢查 Excel
    wb = openpyxl.load_workbook(io.BytesIO(xb), data_only=True)

    sheet = wb["菜單"]
    detail_sheet = wb["採買明細"]
    summary_sheet = wb["採買彙總"]

    assert sheet.max_row == 271
    headers = [sheet.cell(row=1, column=col).value for col in range(1, sheet.max_column + 1)]
    assert headers == [
        "日期",
        "週幾",
        "人數",
        *_expected_role_headers(cfg),
        "成本",
        "目標匹配度",
        "分數拆解(JSON)",
        "分數拆解(易讀)",
    ]
    assert "麵食" in headers
    assert "純蔬" in headers
    assert sheet.cell(row=2, column=3).value == cfg.get("people", 250)
    assert sheet.column_dimensions["A"].width > sheet.column_dimensions["B"].width
    assert detail_sheet.max_row > 1
    assert summary_sheet.max_row > 1
    summary_labels = [summary_sheet.cell(row=r, column=3).value for r in range(2, summary_sheet.max_row + 1)]
    assert "每日小計" in summary_labels
    assert "每週小計" in summary_labels
    assert "全部合計" in summary_labels
    assert summary_sheet.auto_filter.ref == f"A1:I{summary_sheet.max_row}"

    daily_row = next(r for r in range(2, summary_sheet.max_row + 1) if summary_sheet.cell(row=r, column=3).value == "每日小計")
    weekly_row = next(r for r in range(2, summary_sheet.max_row + 1) if summary_sheet.cell(row=r, column=3).value == "每週小計")
    daily_cell = summary_sheet.cell(row=daily_row, column=3)
    weekly_cell = summary_sheet.cell(row=weekly_row, column=3)
    daily_total_cell = summary_sheet.cell(row=daily_row, column=8)
    weekly_total_cell = summary_sheet.cell(row=weekly_row, column=8)

    assert daily_cell.fill.fgColor.rgb == "FFFDE68A"
    assert weekly_cell.fill.fgColor.rgb == "FFBFDBFE"
    assert daily_cell.font.color.rgb == "FF92400E"
    assert weekly_cell.font.color.rgb == "FF1E3A8A"
    assert daily_total_cell.fill.fgColor.rgb == "FFFDE68A"
    assert weekly_total_cell.fill.fgColor.rgb == "FFBFDBFE"
