import io

import openpyxl

from src.menu_planner.api.export_excel import build_plan_workbook


def test_export_excel_uses_dynamic_role_slots_from_config_and_result():
    cfg = {
        "per_day_roles": {"main": 2, "noodle": 1, "side": 4, "veg": 2, "soup": 1, "fruit": 1},
        "per_weekday_roles": {"3": {"main": 1, "noodle": 2, "side": 1, "veg": 3, "soup": 2, "fruit": 1}},
    }
    result = {
        "days": [
            {
                "date": "2026-03-04",
                "items": {
                    "mains": [{"id": "m1", "name": "主菜A"}, {"id": "m2", "name": "主菜B"}],
                    "noodles": [{"id": "n1", "name": "麵食A"}, {"id": "n2", "name": "麵食B"}],
                    "sides": [
                        {"id": "s1", "name": "配菜A"},
                        {"id": "s2", "name": "配菜B"},
                        {"id": "s3", "name": "配菜C"},
                        {"id": "s4", "name": "配菜D"},
                    ],
                    "vegs": [
                        {"id": "v1", "name": "純蔬A"},
                        {"id": "v2", "name": "純蔬B"},
                        {"id": "v3", "name": "純蔬C"},
                    ],
                    "soups": [{"id": "so1", "name": "湯A"}, {"id": "so2", "name": "湯B"}],
                    "fruit": {"id": "f1", "name": "水果A"},
                },
                "day_cost": 123,
                "score_fitness": 88,
                "score_breakdown": {"x": 1},
            }
        ]
    }

    workbook_bytes = build_plan_workbook(cfg, result)
    wb = openpyxl.load_workbook(io.BytesIO(workbook_bytes), data_only=True)
    sheet = wb["菜單"]

    headers = [sheet.cell(row=1, column=col).value for col in range(1, sheet.max_column + 1)]
    assert headers == [
        "日期",
        "週幾",
        "主菜1",
        "主菜2",
        "麵食1",
        "麵食2",
        "配菜1",
        "配菜2",
        "配菜3",
        "配菜4",
        "純蔬1",
        "純蔬2",
        "純蔬3",
        "湯1",
        "湯2",
        "水果",
        "成本",
        "目標匹配度",
        "分數拆解(JSON)",
        "分數拆解(易讀)",
    ]
    values = [sheet.cell(row=2, column=col).value for col in range(1, sheet.max_column + 1)]
    assert values[:16] == [
        "2026-03-04",
        "（三）",
        "主菜A",
        "主菜B",
        "麵食A",
        "麵食B",
        "配菜A",
        "配菜B",
        "配菜C",
        "配菜D",
        "純蔬A",
        "純蔬B",
        "純蔬C",
        "湯A",
        "湯B",
        "水果A",
    ]


def test_export_excel_falls_back_to_single_role_shape_for_legacy_items():
    cfg = {}
    result = {
        "days": [
            {
                "date": "2026-03-05",
                "items": {
                    "main": {"id": "m1", "name": "主菜A"},
                    "sides": [
                        {"id": "s1", "name": "配菜A"},
                        {"id": "s2", "name": "配菜B"},
                    ],
                    "veg": {"id": "v1", "name": "純蔬A"},
                    "soup": {"id": "so1", "name": "湯A"},
                    "fruit": {"id": "f1", "name": "水果A"},
                },
            }
        ]
    }

    workbook_bytes = build_plan_workbook(cfg, result)
    wb = openpyxl.load_workbook(io.BytesIO(workbook_bytes), data_only=True)
    sheet = wb["菜單"]

    headers = [sheet.cell(row=1, column=col).value for col in range(1, sheet.max_column + 1)]
    assert headers[:8] == ["日期", "週幾", "主菜", "配菜1", "配菜2", "純蔬", "湯", "水果"]
    values = [sheet.cell(row=2, column=col).value for col in range(1, 9)]
    assert values == ["2026-03-05", "（四）", "主菜A", "配菜A", "配菜B", "純蔬A", "湯A", "水果A"]


def test_export_excel_adds_weekday_column_and_marks_weekend_offdays():
    cfg = {}
    result = {
        "days": [
            {"date": "2026-03-07", "is_scheduled": False, "items": {}},
            {"date": "2026-03-09", "is_scheduled": True, "items": {}},
        ]
    }

    workbook_bytes = build_plan_workbook(cfg, result)
    wb = openpyxl.load_workbook(io.BytesIO(workbook_bytes), data_only=True)
    sheet = wb["菜單"]

    assert sheet.cell(row=1, column=1).value == "日期"
    assert sheet.cell(row=1, column=2).value == "週幾"
    assert sheet.cell(row=2, column=1).value == "2026-03-07"
    assert sheet.cell(row=2, column=2).value == "（六）"
    assert sheet.cell(row=3, column=2).value == "（一）"
    assert sheet.cell(row=2, column=1).fill.fgColor.rgb == "FFFFE4E6"
    assert sheet.cell(row=2, column=2).fill.fgColor.rgb == "FFFFE4E6"
    assert sheet.cell(row=3, column=1).fill.fgColor.rgb != "FFFFE4E6"


def test_export_excel_human_breakdown_covers_all_dynamic_roles():
    cfg = {"per_day_roles": {"main": 1, "noodle": 1, "side": 1, "veg": 1, "soup": 1, "fruit": 1}}
    result = {
        "days": [
            {
                "date": "2026-03-06",
                "items": {
                    "mains": [{"id": "m1", "name": "主菜A", "inventory_hit_ratio": 0.5}],
                    "noodles": [{"id": "n1", "name": "麵食A", "inventory_hit_ratio": 0.25, "near_expiry_days_min": 3}],
                    "sides": [{"id": "s1", "name": "配菜A", "inventory_hit_ratio": 0.1}],
                    "vegs": [{"id": "v1", "name": "純蔬A", "inventory_hit_ratio": 0.2, "near_expiry_days_min": 6}],
                    "soups": [{"id": "so1", "name": "湯A", "inventory_hit_ratio": 0.3}],
                    "fruits": [{"id": "f1", "name": "水果A", "inventory_hit_ratio": 0.4, "near_expiry_days_min": 2}],
                },
                "score_breakdown": {
                    "near_expiry_bonus": -3,
                    "use_inventory_bonus_others": -2,
                    "use_inventory_bonus_main": -1,
                },
                "score_summary": {"bonus": -6, "penalty": 0, "raw": -6, "fitness": 6},
            }
        ]
    }

    workbook_bytes = build_plan_workbook(cfg, result)
    wb = openpyxl.load_workbook(io.BytesIO(workbook_bytes), data_only=True)
    human = wb["菜單"].cell(row=2, column=wb["菜單"].max_column).value

    assert "主菜庫存命中：主菜A 50%" in human
    assert "非主菜庫存命中：麵食A 25%、配菜A 10%、純蔬A 20%、湯A 30%、水果A 40%" in human
    assert "近到期：麵食A（3天）、純蔬A（6天）、水果A（2天）" in human


def test_export_excel_procurement_detail_uses_role_labels_and_role_order():
    cfg = {}
    result = {
        "days": [
            {
                "date": "2026-03-07",
                "procurement": {
                    "people": 10,
                    "dishes": [
                        {"role": "main", "dish_name": "主菜A", "ingredients": [{"ingredient_name": "雞肉"}]},
                        {"role": "noodle", "dish_name": "麵食A", "ingredients": [{"ingredient_name": "麵"}]},
                        {"role": "veg", "dish_name": "純蔬A", "ingredients": [{"ingredient_name": "青菜"}]},
                    ],
                },
            }
        ]
    }

    workbook_bytes = build_plan_workbook(cfg, result)
    wb = openpyxl.load_workbook(io.BytesIO(workbook_bytes), data_only=True)
    sheet = wb["採買明細"]

    roles = [sheet.cell(row=row, column=2).value for row in range(2, sheet.max_row + 1)]
    assert roles == ["主菜", "麵食", "純蔬"]
